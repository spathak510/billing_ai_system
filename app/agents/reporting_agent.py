# app/agents/reporting_agent.py
from __future__ import annotations

import logging
import os
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

from app.config.settings import settings

logger = logging.getLogger(__name__)

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
ISSUE_FILL = PatternFill("solid", fgColor="FFD966")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
REGION_ORDER = ["AMER", "MEXICO", "AMEA", "EMEAA", "GC"]


class ReportingAgent:
    """Generates formatted Excel reports for billing use-case outputs.

    Outputs produced
    ----------------
    * Zero-data file containing rows whose amount is 0.
    * PAID_CORP_<MONTH_YEAR>.xlsx with regional tabs.
    * PAID_NON-CORP_<MONTH_YEAR>.xlsx with regional tabs.
    """

    def run(
        self,
        corp_df: pd.DataFrame,
        non_corp_df: pd.DataFrame,
        zero_df: pd.DataFrame,
        source_filename: str,
    ) -> dict[str, str]:
        """Write use-case output files and return their absolute paths."""
        os.makedirs(settings.output_dir, exist_ok=True)
        month_year = datetime.now().strftime("%b_%Y").upper()
        stem = Path(source_filename).stem
        zero_path = os.path.join(settings.output_dir, f"ZERO_DATA_{stem}_{month_year}.xlsx")
        corp_path = os.path.join(settings.output_dir, f"PAID_CORP_{month_year}.xlsx")
        non_corp_path = os.path.join(settings.output_dir, f"PAID_NON-CORP_{month_year}.xlsx")

        self._write_zero_file(zero_df, zero_path)
        self._write_paid_file(corp_df, corp_path, user_type_label="CORP")
        self._write_paid_file(non_corp_df, non_corp_path, user_type_label="NON-CORP")

        logger.info("ReportingAgent: generated outputs in %s", settings.output_dir)
        return {
            "zero_data_path": zero_path,
            "corp_paid_path": corp_path,
            "non_corp_paid_path": non_corp_path,
        }

    def _write_zero_file(self, zero_df: pd.DataFrame, output_path: str) -> None:
        wb = Workbook()
        self._write_dataframe(wb, zero_df, sheet_name="ZERO_DATA", highlight_issues=True)
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]
        wb.save(output_path)

    def _write_paid_file(
        self,
        df: pd.DataFrame,
        output_path: str,
        user_type_label: str,
    ) -> None:
        wb = Workbook()

        self._write_summary_sheet(wb, df, user_type_label)
        self._write_dataframe(wb, df, sheet_name="ALL")

        for region in REGION_ORDER:
            region_df = df[df["billing_region"] == region].copy() if "billing_region" in df.columns else pd.DataFrame()
            self._write_dataframe(wb, region_df, sheet_name=region)

        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]
        wb.save(output_path)

    def _write_summary_sheet(self, wb: Workbook, df: pd.DataFrame, user_type_label: str) -> None:
        ws = wb.create_sheet("SUMMARY", 0)
        ws.column_dimensions["A"].width = 34
        ws.column_dimensions["B"].width = 20

        rows = [
            ("User Type", user_type_label),
            ("Total records", len(df)),
            ("Total billed amount", f"{df['amount'].sum():,.2f}" if "amount" in df.columns else "0.00"),
        ]

        for region in REGION_ORDER:
            count = int((df["billing_region"] == region).sum()) if "billing_region" in df.columns else 0
            rows.append((f"{region} records", count))

        ws.append(["Metric", "Value"])
        for cell in ws[1]:
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center")

        for label, value in rows:
            ws.append([label, value])

    def _write_dataframe(
        self,
        wb: Workbook,
        df: pd.DataFrame,
        sheet_name: str,
        highlight_issues: bool = False,
    ) -> None:
        safe_sheet = sheet_name[:31]
        ws = wb.create_sheet(safe_sheet)
        if df.empty:
            ws.append(["No data"])
            return

        for r_idx, row in enumerate(
            dataframe_to_rows(df, index=False, header=True), start=1
        ):
            ws.append(row)
            if r_idx == 1:
                for cell in ws[r_idx]:
                    cell.font = HEADER_FONT
                    cell.fill = HEADER_FILL
                    cell.alignment = Alignment(horizontal="center")
            elif highlight_issues:
                for cell in ws[r_idx]:
                    cell.fill = ISSUE_FILL

        for column in ws.columns:
            max_len = max(
                len(str(cell.value)) if cell.value is not None else 0
                for cell in column
            )
            ws.column_dimensions[column[0].column_letter].width = min(max_len + 4, 50)
