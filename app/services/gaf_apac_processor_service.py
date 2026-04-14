from __future__ import annotations

import logging
from datetime import datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from app.config.settings import settings

logger = logging.getLogger(__name__)


def _default_output_name() -> str:
    return f"GAF_GC_APAC_NON-CORP_{datetime.now().strftime('%B %Y')}"


def _resolve_input_path(input_file_path: str | None) -> Path:
    if input_file_path:
        return Path(input_file_path)

    collection_dir = Path(settings.output_dir) / "APAC" / "APAC_Output"
    collection_files = sorted(
        collection_dir.glob("*_RIR_NONCROP.xlsx"),
        key=lambda p: p.stat().st_mtime,
        reverse=True,
    )
    if collection_files:
        return collection_files[0]

    legacy_cleaned_files = sorted(
        Path(settings.output_dir).glob("cleaned_no_red_*.xlsx"),
        key=lambda p: p.stat().st_mtime,
        reverse=True,
    )
    if legacy_cleaned_files:
        return legacy_cleaned_files[0]

    raise FileNotFoundError("No *_RIR_NONCROP.xlsx file found in output/APAC/APAC_Output.")


def _resolve_template_path(template_path: str | None) -> Path:
    if template_path:
        resolved = Path(template_path)
        if not resolved.exists():
            raise FileNotFoundError(f"Template file not found: {resolved}")
        return resolved

    output_root = Path(settings.output_dir)
    template_dirs = [
        output_root / "APAC" / "GAF_APAC_Processor" / "Template_Format",
        output_root / "APAC" / "GAF_APAC_Processor" / "Template_Formate",
        output_root / "APAC" / "GAF_APAC_Processor" / "Template_formate",
    ]
    for template_dir in template_dirs:
        preferred_template = template_dir / "GAF_GC_APAC_NON-CORP_JANUARY26(updated).xlsx"
        if preferred_template.exists():
            return preferred_template

        template_files = sorted(template_dir.glob("*.xlsx"), key=lambda p: p.stat().st_mtime, reverse=True)
        if template_files:
            return template_files[0]

    legacy_template_dir = output_root / "GAF_APAC_PROCESSER"
    legacy_templates = sorted(legacy_template_dir.glob("*.xlsx"), key=lambda p: p.stat().st_mtime, reverse=True)
    if legacy_templates:
        return legacy_templates[0]

    raise FileNotFoundError(
        "GAF APAC template not found in output/APAC/GAF_APAC_Processor/Template_Format."
    )


def _find_col(columns: list[str], candidates: list[str]) -> str | None:
    normalized = {str(col).strip().upper().replace(" ", ""): col for col in columns}
    for candidate in candidates:
        key = candidate.upper().replace(" ", "")
        if key in normalized:
            return normalized[key]
    return None


def _to_decimal(value: object) -> Decimal:
    text = str(value).strip().replace(",", "")
    if not text:
        return Decimal("0")
    try:
        return Decimal(text)
    except InvalidOperation:
        return Decimal("0")


def _set_cell_value_safe(ws, cell_ref: str, value: object) -> None:
    cell = ws[cell_ref]
    if cell.__class__.__name__ != "MergedCell":
        cell.value = value
        return

    for merged_range in ws.merged_cells.ranges:
        if cell_ref in merged_range:
            ws.cell(row=merged_range.min_row, column=merged_range.min_col).value = value
            return

    ws[cell_ref].value = value


def generate_gaf_apac_output(
    input_file_path: str | None = None,
    template_path: str | None = None,
    output_folder_path: str | None = None,
    submitted_by: str = "",
    output_file_name: str | None = None,
) -> dict[str, str | int | float]:
    """Generate GAF APAC workbook from APAC RIR NONCROP collection data."""
    source_path = _resolve_input_path(input_file_path)
    resolved_template = _resolve_template_path(template_path)
    df = pd.read_excel(source_path, sheet_name=0)

    cols = list(df.columns)
    bu_col = _find_col(cols, ["BU"])
    currency_col = _find_col(cols, ["CURRENCYCODE", "CURRENCY CODE", "CURRENCY"])
    amount_col = _find_col(cols, ["AMOUNT", "BILL_AMOUNT", "BILLING_AMOUNT", "TOTAL_AMOUNT", "NET_AMOUNT", "VALUE"])

    holidex_col = _find_col(cols, ["HOLIDEX", "CUSTID", "CUSTID #"])
    course_col = _find_col(cols, ["COURSE_NAME", "COURSE NAME", "DESCRIPTION"])
    order_col = _find_col(cols, ["ORDER_NO", "ORDER NO", "ORDER_NUMBER", "ORDER NUMBER"])
    offering_date_col = _find_col(cols, ["OFFERING_DATE", "OFFERING DATE"])
    employee_col = _find_col(cols, ["EMPLOYEE", "LEARNERS NAME", "LEARNER", "USERNAME"])

    if not bu_col or not currency_col or not amount_col:
        missing = [
            key
            for key, value in {
                "BU": bu_col,
                "CURRENCYCODE": currency_col,
                "AMOUNT": amount_col,
            }.items()
            if value is None
        ]
        raise ValueError(f"Missing required columns for GAF APAC processing: {', '.join(missing)}")

    apply_revenue_col = None
    for col in cols:
        if "APPLYREVENUE" in str(col).upper().replace(" ", ""):
            apply_revenue_col = col
            break

    gaf_rows: list[dict[str, object]] = df.to_dict(orient="records")

    output_root = (
        Path(output_folder_path)
        if output_folder_path
        else Path(settings.output_dir) / "APAC" / "GAF_APAC_Processor" / "Output"
    )
    output_root.mkdir(parents=True, exist_ok=True)

    workbook = load_workbook(resolved_template)
    if "upload sheet" not in workbook.sheetnames or "GAF" not in workbook.sheetnames:
        raise ValueError("Template must contain 'upload sheet' and 'GAF' sheets")

    upload_sheet = workbook["upload sheet"]
    gaf_sheet = workbook["GAF"]

    # Clear old data rows.
    for row_idx in range(10, max(upload_sheet.max_row, 10000) + 1):
        for col_idx in range(1, 29):
            upload_sheet.cell(row=row_idx, column=col_idx).value = None

    total = Decimal("0")
    row_index = 10
    line_no = 1
    record_count = 0

    for row in gaf_rows:
        amount = _to_decimal(row.get(amount_col, 0))

        upload_sheet.cell(row=row_index, column=1).value = line_no
        upload_sheet.cell(row=row_index, column=2).value = row.get(bu_col, "")
        upload_sheet.cell(row=row_index, column=3).value = row.get(holidex_col, "") if holidex_col else ""
        current_dt = datetime.now()
        upload_sheet.cell(row=row_index, column=5).value = f"{current_dt.month}/{current_dt.day}/{current_dt.year}"
        upload_sheet.cell(row=row_index, column=6).value = "LMS Training"
        upload_sheet.cell(row=row_index, column=7).value = row.get(course_col, "") if course_col else ""
        upload_sheet.cell(row=row_index, column=8).value = row.get(currency_col, "")
        upload_sheet.cell(row=row_index, column=9).value = float(amount)
        upload_sheet.cell(row=row_index, column=11).value = row.get(apply_revenue_col, "") if apply_revenue_col else ""
        upload_sheet.cell(row=row_index, column=15).value = row.get(course_col, "") if course_col else ""
        upload_sheet.cell(row=row_index, column=24).value = ""
        upload_sheet.cell(row=row_index, column=25).value = ""
        upload_sheet.cell(row=row_index, column=26).value = row.get(order_col, "") if order_col else ""
        upload_sheet.cell(row=row_index, column=27).value = row.get(offering_date_col, "") if offering_date_col else ""
        upload_sheet.cell(row=row_index, column=28).value = row.get(employee_col, "") if employee_col else ""

        total += amount
        record_count += 1
        row_index += 1
        line_no += 1

    gaf_sheet["G4"] = datetime.now().strftime("%B %d, %Y")
    gaf_sheet["G5"] = submitted_by
    gaf_sheet["M26"] = float(total)

    _set_cell_value_safe(upload_sheet, "I5", record_count)

    base_name = (output_file_name or "GAF_GC_APAC_NON-CORP").strip() or "GAF_GC_APAC_NON-CORP"
    if base_name.lower().endswith(".xlsx"):
        base_name = base_name[:-5]
    month_year = datetime.now().strftime("%B %Y")
    if not base_name.endswith(month_year):
        base_name = f"{base_name}_{month_year}"
    final_file_name = f"{base_name}.xlsx"
    output_path = output_root / final_file_name

    if output_path.exists():
        output_path.unlink()

    workbook.save(output_path)

    logger.info("Generated GAF APAC output: %s", output_path)
    return {
        "gaf_apac_output_path": str(output_path),
        "gaf_apac_records": record_count,
        "gaf_apac_total": float(total),
        "template_file": str(resolved_template),
        "source_file": str(source_path),
    }
