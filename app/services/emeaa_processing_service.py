from __future__ import annotations

import logging
from datetime import datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from app.config.settings import settings

logger = logging.getLogger(__name__)


def _resolve_input_path(input_file_path: str | None) -> Path:
    if input_file_path:
        return Path(input_file_path)

    output_dir = Path(settings.output_dir)
    cleaned_files = sorted(output_dir.glob("cleaned_no_red_*.xlsx"), key=lambda p: p.stat().st_mtime, reverse=True)
    if cleaned_files:
        return cleaned_files[0]

    raise FileNotFoundError("No cleaned_no_red_*.xlsx file found in output folder.")


def _to_decimal(value: object) -> Decimal:
    text = str(value).strip().replace(",", "")
    if not text:
        return Decimal("0")
    try:
        return Decimal(text)
    except InvalidOperation:
        return Decimal("0")


def _generate_emeaa_noncorp_formatted_output(
    rows: list[dict[str, object]],
    source_columns: list[str],
) -> str | None:
    base_dir = Path(settings.output_dir) / "EMEAA"
    template_path = base_dir / "EMEAA_INP_FORMAT" / "EMEAA_Intercompany billing lines_January26.xlsx"
    output_dir = base_dir / "EMEAA_Output"
    output_dir.mkdir(parents=True, exist_ok=True)

    if not template_path.exists():
        logger.warning("EMEAA template not found: %s", template_path)
        return None

    wb = load_workbook(template_path)
    if "BILLING LINES" not in wb.sheetnames:
        logger.warning("EMEAA template missing BILLING LINES sheet: %s", template_path)
        return None

    ws = wb["BILLING LINES"]
    headers = [cell.value for cell in ws[1] if cell.value not in (None, "")]
    header_names = [str(h).strip() for h in headers]
    if not header_names:
        logger.warning("No BILLING LINES headers found in EMEAA template: %s", template_path)
        return None

    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)

    source_map = {str(col).strip().upper(): col for col in source_columns}
    for row_dict in rows:
        output_row: list[object] = []
        for header in header_names:
            source_col = source_map.get(header.upper())
            output_row.append(row_dict.get(source_col) if source_col else None)
        ws.append(output_row)

    if "RIR" in wb.sheetnames:
        wb["RIR"]["F10"] = datetime.now().date()

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = output_dir / f"EMEAA_NON_CROP_{timestamp}.xlsx"
    wb.save(output_path)
    logger.info("Generated EMEAA formatted output: %s", output_path)
    return str(output_path)


def generate_emeaa_processing_output(input_file_path: str | None = None) -> dict[str, str | int]:
    """Run EMEAA V1/V2/GAF logic from cleaned data and produce formatted NON_CROP output."""
    source_path = _resolve_input_path(input_file_path)
    df = pd.read_excel(source_path, sheet_name=0)

    col_map = {str(col).upper(): col for col in df.columns}
    bu_col = col_map.get("BU")
    currency_col = col_map.get("CURRENCYCODE")
    amount_col = col_map.get("AMOUNT")
    user_type_col = col_map.get("USER_TYPE")

    if not bu_col or not currency_col or not amount_col or not user_type_col:
        missing = [
            key
            for key, col in {
                "BU": bu_col,
                "CURRENCYCODE": currency_col,
                "AMOUNT": amount_col,
                "USER_TYPE": user_type_col,
            }.items()
            if col is None
        ]
        raise ValueError(f"Missing required columns for EMEAA processing: {', '.join(missing)}")

    emeaa_v1_rows: list[dict[str, object]] = []
    emeaa_v2_rows: list[dict[str, object]] = []
    emeaa_gaf_rows: list[dict[str, object]] = []

    for _, row in df.iterrows():
        bu = str(row.get(bu_col, "")).strip().upper()
        if not bu.startswith("H"):
            continue

        currency = str(row.get(currency_col, "")).strip().upper()
        amount = _to_decimal(row.get(amount_col, "0"))
        user_type = str(row.get(user_type_col, "")).strip().upper()

        if currency == "EUR":
            amount = amount / Decimal("0.86")
            currency = "USD"

        base_row = row.to_dict()
        base_row[amount_col] = float(amount)
        base_row[currency_col] = currency

        emeaa_v1_rows.append(dict(base_row))

        v2_row = dict(base_row)
        if amount >= 0:
            v2_row["INVOICE_NO"] = "N/A"
            v2_row["INVOICE_DATE"] = "N/A"
        emeaa_v2_rows.append(v2_row)

        if amount < 0 and user_type != "C":
            emeaa_gaf_rows.append(dict(base_row))

    emeaa_noncorp_rows = [
        row for row in emeaa_v2_rows if str(row.get(user_type_col, "")).strip().upper() != "C"
    ]

    formatted_path = _generate_emeaa_noncorp_formatted_output(
        rows=emeaa_noncorp_rows,
        source_columns=list(df.columns),
    )

    result: dict[str, str | int] = {
        "emeaa_v1_rows": len(emeaa_v1_rows),
        "emeaa_v2_rows": len(emeaa_v2_rows),
        "emeaa_gaf_rows": len(emeaa_gaf_rows),
        "emeaa_noncorp_rows": len(emeaa_noncorp_rows),
    }
    if formatted_path:
        result["emeaa_noncorp_formatted_path"] = formatted_path

    return result
