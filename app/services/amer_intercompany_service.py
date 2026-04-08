from __future__ import annotations

import calendar
import logging
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook, load_workbook

from app.config.settings import settings

logger = logging.getLogger(__name__)


def _normalized_key(value: object) -> str:
    return str(value).strip().upper().replace(" ", "")


def _resolve_input_path(input_file_path: str | None) -> Path:
    if input_file_path:
        return Path(input_file_path)

    output_dir = Path(settings.output_dir)
    amer_files = sorted(output_dir.glob("AMER_*.xlsx"), key=lambda p: p.stat().st_mtime, reverse=True)
    if amer_files:
        return amer_files[0]

    cleaned_files = sorted(output_dir.glob("cleaned_no_red_*.xlsx"), key=lambda p: p.stat().st_mtime, reverse=True)
    if cleaned_files:
        return cleaned_files[0]

    raise FileNotFoundError("No AMER_*.xlsx or cleaned_no_red_*.xlsx file found in output folder.")


def _resolve_template_path(template_path: str | None) -> Path | None:
    if template_path:
        resolved = Path(template_path)
        if not resolved.exists():
            raise FileNotFoundError(f"Template file not found: {resolved}")
        return resolved

    return None


def _default_output_name() -> str:
    month = calendar.month_name[datetime.now().month]
    year = datetime.now().year
    return f"AMER_Intercompany billing lines_{month} {year}.xlsx"


def _next_available_path(path: Path) -> Path:
    if not path.exists():
        return path

    stem = path.stem
    suffix = path.suffix
    index = 1
    while True:
        candidate = path.parent / f"{stem}_{index}{suffix}"
        if not candidate.exists():
            return candidate
        index += 1


def _create_default_workbook() -> Workbook:
    wb = Workbook()
    rir_sheet = wb.active
    rir_sheet.title = "RIR"
    billing_sheet = wb.create_sheet(title="BILLING LINES")

    billing_headers = [
        "USERNAME",
        "EMPLOYEE",
        "HOLIDEX",
        "AMOUNT",
        "CURRENCYCODE",
        "COST_CENTER",
        "ORDER_NO",
        "COURSE_NAME",
        "FACILITY",
        "OFFERING_ID",
        "INSTRUCTOR",
        "OFFERING_DATE",
        "COUNTRY",
        "REGION",
        "USER_TYPE",
        "TRANSTYPECODE",
        "PAY_DATE",
        "NAME",
        "DELIVERED_ON",
        "REVENUE",
        "BU",
    ]
    for col_idx, header in enumerate(billing_headers, start=1):
        billing_sheet.cell(row=1, column=col_idx).value = header

    return wb


def _load_workbook_for_output(template_path: Path | None) -> tuple[Workbook, str]:
    if template_path is None:
        return _create_default_workbook(), "generated"

    wb = load_workbook(template_path)
    if "RIR" in wb.sheetnames and "BILLING LINES" in wb.sheetnames:
        return wb, str(template_path)

    logger.warning(
        "AMER Intercompany template %s does not contain RIR and BILLING LINES sheets. Using generated workbook.",
        template_path,
    )
    return _create_default_workbook(), "generated"


def _find_revenue_column(columns: list[object]) -> str | None:
    for col in columns:
        if "REVEN" in _normalized_key(col):
            return str(col)
    return None


def _value_from_row(row_dict: dict[str, object], column_name: str) -> object:
    target = _normalized_key(column_name)
    for key, value in row_dict.items():
        if _normalized_key(key) == target:
            return value
    return ""


def _filter_amer_rows(df: pd.DataFrame) -> pd.DataFrame:
    col_map = {_normalized_key(col): col for col in df.columns}
    region_col = col_map.get("REGION")
    bu_col = col_map.get("BU")

    if region_col is not None:
        region_series = df[region_col].fillna("").astype(str).str.strip().str.upper()
        amer_df = df[region_series.eq("AMER")].copy()
    elif bu_col is not None:
        bu_series = df[bu_col].fillna("").astype(str).str.strip().str.upper()
        amer_df = df[bu_series.str.startswith("A")].copy()
    else:
        amer_df = df.copy()

    return amer_df


def _clear_billing_lines(ws) -> None:
    if ws.max_row <= 1:
        return
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=26):
        for cell in row:
            cell.value = None


def generate_amer_intercompany_output(
    input_file_path: str | None = None,
    template_path: str | None = None,
    output_folder_path: str | None = None,
    base_file_name: str | None = None,
    request_date: str | None = None,
    request_name: str = "",
    account_number: str = "",
) -> dict[str, str | int]:
    """Generate AMER Intercompany billing-lines workbook from cleaned/split AMER data."""
    source_path = _resolve_input_path(input_file_path)
    resolved_template = _resolve_template_path(template_path)

    target_dir = Path(output_folder_path) if output_folder_path else Path(settings.output_dir) / "AMER_Intercompny"
    target_dir.mkdir(parents=True, exist_ok=True)

    final_file_name = (base_file_name or _default_output_name()).strip()
    if not final_file_name.lower().endswith(".xlsx"):
        final_file_name += ".xlsx"

    output_path = target_dir / final_file_name
    if resolved_template is not None and output_path.resolve() == resolved_template.resolve():
        output_path = _next_available_path(output_path)
    elif output_path.exists():
        output_path = _next_available_path(output_path)

    df = pd.read_excel(source_path, sheet_name=0)
    amer_df = _filter_amer_rows(df)
    if amer_df.empty:
        raise ValueError("BillingCollection is EMPTY for AMER Intercompany processing.")

    wb, template_used = _load_workbook_for_output(resolved_template)

    rir_sheet = wb["RIR"]
    billing_sheet = wb["BILLING LINES"]

    rir_sheet["F10"] = request_date or datetime.now().strftime("%Y-%m-%d")
    rir_sheet["P10"] = request_name
    rir_sheet["F11"] = account_number

    _clear_billing_lines(billing_sheet)

    rows = amer_df.to_dict(orient="records")
    revenue_col = _find_revenue_column(list(amer_df.columns))

    write_row = 2
    for row_dict in rows:
        billing_sheet.cell(write_row, 1).value = _value_from_row(row_dict, "USERNAME")
        billing_sheet.cell(write_row, 2).value = _value_from_row(row_dict, "EMPLOYEE")
        billing_sheet.cell(write_row, 3).value = _value_from_row(row_dict, "HOLIDEX")
        billing_sheet.cell(write_row, 4).value = _value_from_row(row_dict, "AMOUNT")
        billing_sheet.cell(write_row, 5).value = _value_from_row(row_dict, "CURRENCYCODE")
        billing_sheet.cell(write_row, 6).value = _value_from_row(row_dict, "COST_CENTER")
        billing_sheet.cell(write_row, 7).value = _value_from_row(row_dict, "ORDER_NO")
        billing_sheet.cell(write_row, 8).value = _value_from_row(row_dict, "COURSE_NAME")
        billing_sheet.cell(write_row, 9).value = _value_from_row(row_dict, "FACILITY")
        billing_sheet.cell(write_row, 10).value = _value_from_row(row_dict, "OFFERING_ID")
        billing_sheet.cell(write_row, 11).value = _value_from_row(row_dict, "INSTRUCTOR")
        billing_sheet.cell(write_row, 12).value = _value_from_row(row_dict, "OFFERING_DATE")
        billing_sheet.cell(write_row, 13).value = _value_from_row(row_dict, "COUNTRY")
        billing_sheet.cell(write_row, 14).value = _value_from_row(row_dict, "REGION")
        billing_sheet.cell(write_row, 15).value = _value_from_row(row_dict, "USER_TYPE")
        billing_sheet.cell(write_row, 16).value = _value_from_row(row_dict, "TRANSTYPECODE")
        billing_sheet.cell(write_row, 17).value = _value_from_row(row_dict, "PAY_DATE")
        billing_sheet.cell(write_row, 18).value = _value_from_row(row_dict, "NAME")
        billing_sheet.cell(write_row, 19).value = _value_from_row(row_dict, "DELIVERED_ON")
        billing_sheet.cell(write_row, 20).value = row_dict.get(revenue_col, "") if revenue_col else ""
        billing_sheet.cell(write_row, 21).value = _value_from_row(row_dict, "BU")
        write_row += 1

    wb.save(output_path)
    logger.info("Generated AMER Intercompany output: %s", output_path)
    return {
        "amer_intercompany_file": str(output_path),
        "amer_intercompany_rows": len(rows),
        "template_file": template_used,
        "source_file": str(source_path),
    }