from __future__ import annotations

import calendar
import logging
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from app.config.settings import settings

logger = logging.getLogger(__name__)


def _normalized_key(value: object) -> str:
    return str(value).strip().upper().replace(" ", "")


def _resolve_input_path(input_file_path: str | None) -> Path:
    if input_file_path:
        return Path(input_file_path)

    output_dir = Path(settings.output_dir) / "EMEAA" / "Output"
    emeaa_v2_files = sorted(output_dir.glob("EMEAA_V2.xlsx"), key=lambda p: p.stat().st_mtime, reverse=True)
    if emeaa_v2_files:
        return emeaa_v2_files[0]

    raise FileNotFoundError("No EMEAA_V2.xlsx file found in output/EMEAA/Output.")


def _resolve_template_path(template_path: str | None) -> Path:
    if template_path:
        resolved = Path(template_path)
        if not resolved.exists():
            raise FileNotFoundError(f"Template file not found: {resolved}")
        return resolved

    template_dir = Path(settings.output_dir) / "EMEAA" / "EMEAA_Intercompany" / "Template_Formate"
    template_files = sorted(template_dir.glob("*.xlsx"), key=lambda p: p.stat().st_mtime, reverse=True)
    if template_files:
        return template_files[0]

    raise FileNotFoundError(
        "EMEAA Intercompany template not found in output/EMEAA/EMEAA_Intercompany/Template_Formate."
    )


def _default_output_name() -> str:
    month = calendar.month_name[datetime.now().month]
    year = datetime.now().year
    return f"EMEAA_Intercompany billing lines_{month} {year}.xlsx"


def _value_from_row(row_dict: dict[str, object], column_name: str) -> object:
    target = _normalized_key(column_name)
    for key, value in row_dict.items():
        if _normalized_key(key) == target:
            return value
    return ""


def _clear_billing_lines(ws) -> None:
    if ws.max_row <= 1:
        return

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=26):
        for cell in row:
            cell.value = None


def _find_revenue_column(columns: list[object]) -> str | None:
    for col in columns:
        if "REVEN" in _normalized_key(col):
            return str(col)
    return None


def generate_emeaa_intercompany_output(
    input_file_path: str | None = None,
    template_path: str | None = None,
    output_folder_path: str | None = None,
    base_file_name: str | None = None,
    request_date: str | None = None,
    request_name: str = "",
    account_number: str = "",
) -> dict[str, str | int]:
    """Generate EMEAA Intercompany workbook from EMEAA V2 collection data."""
    source_path = _resolve_input_path(input_file_path)
    resolved_template = _resolve_template_path(template_path)

    target_dir = (
        Path(output_folder_path)
        if output_folder_path
        else Path(settings.output_dir) / "EMEAA" / "EMEAA_Intercompany" / "Output"
    )
    target_dir.mkdir(parents=True, exist_ok=True)

    final_file_name = (base_file_name or _default_output_name()).strip()
    if not final_file_name.lower().endswith(".xlsx"):
        final_file_name += ".xlsx"
    output_path = target_dir / final_file_name

    df = pd.read_excel(source_path, sheet_name=0)
    if df.empty:
        raise ValueError("BillingCollection is EMPTY for EMEAA Intercompany processing.")

    wb = load_workbook(resolved_template)
    if "RIR" not in wb.sheetnames or "BILLING LINES" not in wb.sheetnames:
        raise ValueError(
            f"EMEAA Intercompany template {resolved_template} does not contain RIR and BILLING LINES sheets."
        )

    rir_sheet = wb["RIR"]
    billing_sheet = wb["BILLING LINES"]

    rir_sheet["F10"] = request_date or datetime.now().strftime("%Y-%m-%d")
    rir_sheet["P10"] = request_name
    rir_sheet["F11"] = account_number

    _clear_billing_lines(billing_sheet)

    rows = df.to_dict(orient="records")
    revenue_col = _find_revenue_column(list(df.columns))

    write_row = 2
    for row_dict in rows:
        billing_sheet.cell(write_row, 1).value = _value_from_row(row_dict, "USERNAME")
        billing_sheet.cell(write_row, 2).value = _value_from_row(row_dict, "EMPLOYEE")
        billing_sheet.cell(write_row, 3).value = _value_from_row(row_dict, "HOLIDEX")
        billing_sheet.cell(write_row, 4).value = _value_from_row(row_dict, "AMOUNT")
        billing_sheet.cell(write_row, 5).value = _value_from_row(row_dict, "CURRENCYCODE")
        billing_sheet.cell(write_row, 6).value = _value_from_row(row_dict, "COST_CENTER")
        billing_sheet.cell(write_row, 7).value = _value_from_row(row_dict, "ORDER_NO")
        billing_sheet.cell(write_row, 8).value = _value_from_row(row_dict, "COURSE_NAME")
        billing_sheet.cell(write_row, 9).value = _value_from_row(row_dict, "FACILITY")
        billing_sheet.cell(write_row, 10).value = _value_from_row(row_dict, "OFFERING_ID")
        billing_sheet.cell(write_row, 11).value = _value_from_row(row_dict, "INSTRUCTOR")
        billing_sheet.cell(write_row, 12).value = _value_from_row(row_dict, "OFFERING_DATE")
        billing_sheet.cell(write_row, 13).value = _value_from_row(row_dict, "COUNTRY")
        billing_sheet.cell(write_row, 14).value = _value_from_row(row_dict, "REGION")
        billing_sheet.cell(write_row, 15).value = _value_from_row(row_dict, "USER_TYPE")
        billing_sheet.cell(write_row, 16).value = _value_from_row(row_dict, "TRANSTYPECODE")
        billing_sheet.cell(write_row, 17).value = _value_from_row(row_dict, "PAY_DATE")
        billing_sheet.cell(write_row, 18).value = _value_from_row(row_dict, "NAME")
        billing_sheet.cell(write_row, 19).value = _value_from_row(row_dict, "DELIVERED_ON")
        billing_sheet.cell(write_row, 20).value = row_dict.get(revenue_col, "") if revenue_col else ""
        billing_sheet.cell(write_row, 21).value = _value_from_row(row_dict, "BU")
        write_row += 1

    if output_path.exists():
        output_path.unlink()

    wb.save(output_path)
    logger.info("Generated EMEAA Intercompany output: %s", output_path)
    return {
        "emeaa_intercompany_file": str(output_path),
        "emeaa_intercompany_rows": len(rows),
        "template_file": str(resolved_template),
        "source_file": str(source_path),
    }
