from openpyxl import load_workbook, Workbook
from pathlib import Path
import logging

def generate_input_emeaa_intercompany_result(input_file_path: str) -> str:
    """
    Generate EMEAA_INTERCOMPANY Excel file from APAC_GC_Collection (input_file_path).
    Only rows where REGION contains 'AMEA', 'APAC', or 'GC', USER_TYPE not 'C', and BU starts with 'H' are included.
    Returns the output file path.
    """
    logger = logging.getLogger(__name__)
    input_path = Path(input_file_path)
    from app.config.settings import settings
    output_dir = Path(settings.output_dir) / "EMEAA" / "EMEAA_Intercompany" / "Input"
    output_dir.mkdir(parents=True, exist_ok=True)
    output_name = f"EMEAA_INTERCOMPANY_{input_path.stem}.xlsx"
    output_path = output_dir / output_name

    wb = load_workbook(input_path)
    ws = wb.active
    rows = list(ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column, values_only=True))
    if not rows:
        logger.warning("No data found in input file: %s", input_file_path)
        return str(output_path)
    header = list(rows[0])
    col_idx = {str(h).strip().upper(): i for i, h in enumerate(header)}

    # Create output workbook
    out_wb = Workbook()
    out_ws = out_wb.active
    out_ws.title = "EMEAA_INTERCOMPANY"
    out_ws.append(header)

    for row in rows[1:]:
        bu = str(row[col_idx.get("BU", -1)]).strip().upper() if col_idx.get("BU") is not None and row[col_idx.get("BU")] is not None else ""
        user_type = str(row[col_idx.get("USER_TYPE", -1)]).strip().upper() if col_idx.get("USER_TYPE") is not None and row[col_idx.get("USER_TYPE")] is not None else ""
        region = str(row[col_idx.get("REGION", -1)]).strip().upper() if col_idx.get("REGION") is not None and row[col_idx.get("REGION")] is not None else ""
        # FINAL FILTER LOGIC
        if bu != "" and user_type != "" and region != "":
            if ("AMEA" in region or "APAC" in region or "GC" in region) and user_type != "C" and bu.startswith("H"):
                out_ws.append(row)

    out_wb.save(output_path)
    logger.info("Created EMEAA_INTERCOMPANY file: %s", output_path)
    return str(output_path)
