from openpyxl import load_workbook, Workbook
from pathlib import Path
import logging

def generate_input_apac_gc_intewrcompany(input_file_path: str) -> str:
    """
    Generate APAC_GC_INTERCOMPANY Excel file from APAC_GC_Collection (input_file_path).
    Only rows where REGION == 'EMEAA', USER_TYPE not 'C' and not empty, and BU starts with 'P' are included.
    Returns the output file path.
    """
    logger = logging.getLogger(__name__)
    from app.config.settings import settings
    input_path = Path(input_file_path)
    output_dir = Path(settings.output_dir) / "APAC" / "APAC_Intercompny" / "Input"
    output_dir.mkdir(parents=True, exist_ok=True)
    output_name = f"APAC_GC_INTERCOMPANY_{input_path.stem}.xlsx"
    output_path = output_dir / output_name

    wb = load_workbook(input_path)
    ws = wb.active
    rows = list(ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column, values_only=True))
    if not rows:
        logger.warning("No data found in input file: %s", input_file_path)
        return str(output_path)
    header = list(rows[0])
    col_idx = {str(h).strip().upper(): i for i, h in enumerate(header)}

    # Create output workbook
    out_wb = Workbook()
    out_ws = out_wb.active
    out_ws.title = "APAC_GC_INTERCOMPANY"
    out_ws.append(header)

    for row in rows[1:]:
        bu = str(row[col_idx.get("BU", -1)]).strip().upper() if col_idx.get("BU") is not None else ""
        user_type = str(row[col_idx.get("USER_TYPE", -1)]).strip().upper() if col_idx.get("USER_TYPE") is not None else ""
        region = str(row[col_idx.get("REGION", -1)]).strip().upper() if col_idx.get("REGION") is not None else ""
        # FINAL FILTER: REGION == 'EMEAA', USER_TYPE not 'C' and not empty, BU starts with 'P'
        if region == "EMEAA":
            if user_type != "C" and user_type != "":
                if bu.startswith("P"):
                    out_ws.append(row)

    out_wb.save(output_path)
    logger.info("Created APAC_GC_INTERCOMPANY file: %s", output_path)
    return str(output_path)
