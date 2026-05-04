from __future__ import annotations

import logging
from pathlib import Path

from openpyxl import Workbook, load_workbook

from app.config.settings import settings
from app.services.apac_processing_service import (
    generate_apac_gc_intewrcompany_output,
    generate_apac_processing_output,
)
from app.services.amer_intercompany_service import generate_amer_intercompany_output
from app.services.emeaa_intercompany_service import generate_emeaa_intercompany_output
from app.services.emeaa_processing_service import generate_emeaa_processing_output
from app.services.gaf_apac_processor_service import generate_gaf_apac_output
from app.services.rir_apac_processor_service import generate_rir_apac_output
from app.services.jrf_processor_service import generate_jrf_output
from app.services.peoplesoft_output_service import generate_amer_peoplesoft_output
from app.services.apac_gc_intewrcompany_input_service import generate_input_apac_gc_intewrcompany
from app.services.emeaa_intercompany_input_service import generate_input_emeaa_intercompany_result

logger = logging.getLogger(__name__)


def _ensure_expected_output_dirs() -> None:
    output_root = Path(settings.output_dir)
    expected_dirs = [
        output_root / "Region_Wise_Split",
        output_root / "AMER" / "AMER_Output",
        output_root / "AMER_Intercompny" / "Output",
        output_root / "APAC" / "APAC_Output",
        output_root / "APAC" / "APAC_Intercompny" / "Output",
        output_root / "APAC" / "GAF_APAC_Processor" / "Output",
        output_root / "APAC" / "APAC_GC_RIR" / "Output",
        output_root / "EMEAA" / "Output",
        output_root / "EMEAA" / "EMEAA_Intercompany" / "Output",
        output_root / "JRF" / "Template_Formate",
        output_root / "JRF" / "Output",
    ]
    for directory in expected_dirs:
        directory.mkdir(parents=True, exist_ok=True)


def _find_column(header_row: list[object], column_name: str) -> int | None:
    for idx, value in enumerate(header_row):
        if str(value).strip().upper() == column_name:
            return idx
    return None


def _split_user_type_collections(cleaned_workbook: Workbook, source_stem: str) -> None:
    from copy import deepcopy
    import re
    output_dir = Path(settings.output_dir) / "Corp_NonCorp_Split"
    output_dir.mkdir(parents=True, exist_ok=True)

    # Helper to create a new workbook with a single sheet and header
    def create_collection(header, title):
        wb = Workbook()
        ws = wb.active
        ws.title = title
        ws.append(header)
        return wb, ws

    # We'll process only the first sheet for simplicity (can be extended for multi-sheet)
    cleaned_sheet = cleaned_workbook.worksheets[0]
    rows = list(
        cleaned_sheet.iter_rows(
            min_row=1,
            max_row=cleaned_sheet.max_row,
            max_col=cleaned_sheet.max_column,
            values_only=False,  # Need cell objects for cleaning
        )
    )
    if not rows:
        return
    header = [cell.value for cell in rows[0]]
    col_idx = {str(h).strip().upper(): i for i, h in enumerate(header)}

    # Create all collections (workbooks and sheets)
    def new_ws(name):
        wb = Workbook()
        ws = wb.active
        ws.title = name
        ws.append(header)
        return wb, ws
    # Main collections
    ZeroCollection_wb, ZeroCollection = new_ws("ZeroCollection")
    CorpCollection_wb, CorpCollection = new_ws("CorpCollection")
    NonCorpCollection_wb, NonCorpCollection = new_ws("NonCorpCollection")
    AMERCollection_wb, AMERCollection = new_ws("AMERCollection")
    MexicoCollection_wb, MexicoCollection = new_ws("MexicoCollection")
    APACCollection_wb, APACCollection = new_ws("APACCollection")
    EMEAACollection_wb, EMEAACollection = new_ws("EMEAACollection")
    GCCollection_wb, GCCollection = new_ws("GCCollection")
    # Crop/NonCrop by region
    AMER_CROP_wb, AMER_CROP = new_ws("AMER_CROP")
    AMER_NONCROP_wb, AMER_NONCROP = new_ws("AMER_NONCROP")
    MEXICO_CROP_wb, MEXICO_CROP = new_ws("MEXICO_CROP")
    MEXICO_NONCROP_wb, MEXICO_NONCROP = new_ws("MEXICO_NONCROP")
    APAC_CROP_wb, APAC_CROP = new_ws("APAC_CROP")
    APAC_NONCROP_wb, APAC_NONCROP = new_ws("APAC_NONCROP")
    EMEAA_CROP_wb, EMEAA_CROP = new_ws("EMEAA_CROP")
    EMEAA_NONCROP_wb, EMEAA_NONCROP = new_ws("EMEAA_NONCROP")
    GC_CROP_wb, GC_CROP = new_ws("GC_CROP")
    GC_NONCROP_wb, GC_NONCROP = new_ws("GC_NONCROP")
    # Combined
    APAC_AMEA_COMBINED_wb, APAC_AMEA_COMBINED = new_ws("APAC_AMEA_COMBINED")
    APAC_AMEA_CROP_wb, APAC_AMEA_CROP = new_ws("APAC_AMEA_CROP")
    APAC_AMEA_NONCROP_wb, APAC_AMEA_NONCROP = new_ws("APAC_AMEA_NONCROP")
    APAC_AMEA_GC_COMBINED_wb, APAC_AMEA_GC_COMBINED = new_ws("APAC_AMEA_GC_COMBINED")
    APAC_AMEA_GC_CROP_wb, APAC_AMEA_GC_CROP = new_ws("APAC_AMEA_GC_CROP")
    APAC_AMEA_GC_NONCROP_wb, APAC_AMEA_GC_NONCROP = new_ws("APAC_AMEA_GC_NONCROP")

    # Helper to get value by column name
    def get(row, name):
        idx = col_idx.get(name)
        if idx is not None and idx < len(row):
            return row[idx].value if hasattr(row[idx], "value") else row[idx]
        return ""

    # Main row processing
    for row in rows[1:]:
        row_values = [cell.value for cell in row]
        # --- SAFE AMOUNT ---
        amount = 0
        amount_val = get(row, "AMOUNT")
        try:
            if amount_val is not None and str(amount_val).strip() != "":
                amount = float(amount_val)
        except Exception:
            amount = 0
        if amount == 0:
            ZeroCollection.append(row_values)
            continue
        # --- USER TYPE ---
        user_type = str(get(row, "USER_TYPE")).strip().upper()
        if user_type == "C":
            CorpCollection.append(row_values)
        elif user_type in ("F", "H"):
            NonCorpCollection.append(row_values)
        # --- REGION LOGIC ---
        region = str(get(row, "REGION")).strip().upper()
        country = str(get(row, "COUNTRY")).strip().upper()
        # AMER / GLOBAL
        if "GLOBAL" in region or region == "AMER":
            if country in ("UNITED STATES", "CANADA"):
                AMERCollection.append(row_values)
                if user_type == "C":
                    AMER_CROP.append(row_values)
                else:
                    AMER_NONCROP.append(row_values)
            else:
                MexicoCollection.append(row_values)
                if user_type == "C":
                    MEXICO_CROP.append(row_values)
                else:
                    MEXICO_NONCROP.append(row_values)
        # APAC (AMEA + APAC)
        elif region in ("AMEA", "APAC"):
            APACCollection.append(row_values)
            APAC_AMEA_COMBINED.append(row_values)
            if user_type == "C":
                APAC_CROP.append(row_values)
                APAC_AMEA_CROP.append(row_values)
            else:
                APAC_NONCROP.append(row_values)
                APAC_AMEA_NONCROP.append(row_values)
        # EMEAA
        elif region == "EMEAA":
            EMEAACollection.append(row_values)
            if user_type == "C":
                EMEAA_CROP.append(row_values)
            else:
                EMEAA_NONCROP.append(row_values)
        # GC
        elif region == "GC":
            GCCollection.append(row_values)
            if user_type == "C":
                GC_CROP.append(row_values)
            else:
                GC_NONCROP.append(row_values)
        # --- HOLIDEX FIX ---
        holidexC = str(get(row, "HOLIDEX")).strip()
        holidexY = str(get(row, "PERSON_HOLIDEX")).strip()
        if len(holidexC) != 5 and len(holidexY) == 5:
            idx = col_idx.get("HOLIDEX")
            if idx is not None and idx < len(row):
                row[idx].value = holidexY
        # --- CLEAN COURSE NAME ---
        course_name = str(get(row, "COURSE_NAME"))
        course_name_clean = re.sub(r"[^a-zA-Z0-9\-\&\(\) ]", "", course_name)
        idx = col_idx.get("COURSE_NAME")
        if idx is not None and idx < len(row):
            row[idx].value = course_name_clean

    # =============================
    # FINAL COMBINE (APAC_AMEA + GC)
    # =============================
    # Combine APAC_AMEA
    for row in APAC_AMEA_COMBINED.iter_rows(min_row=2, max_row=APAC_AMEA_COMBINED.max_row, values_only=True):
        APAC_AMEA_GC_COMBINED.append(row)
        user_type = str(row[col_idx.get("USER_TYPE", -1)]).strip().upper() if col_idx.get("USER_TYPE") is not None else ""
        if user_type == "C":
            APAC_AMEA_GC_CROP.append(row)
        else:
            APAC_AMEA_GC_NONCROP.append(row)
    # Combine GC
    for row in GCCollection.iter_rows(min_row=2, max_row=GCCollection.max_row, values_only=True):
        APAC_AMEA_GC_COMBINED.append(row)
        user_type = str(row[col_idx.get("USER_TYPE", -1)]).strip().upper() if col_idx.get("USER_TYPE") is not None else ""
        if user_type == "C":
            APAC_AMEA_GC_CROP.append(row)
        else:
            APAC_AMEA_GC_NONCROP.append(row)

    # Save all collections to files
    def save_wb(wb, name):
        path = _next_available_path(output_dir / f"{name}_{source_stem}.xlsx")
        wb.save(path)
        logger.info("Created %s file: %s", name, path)
    save_wb(ZeroCollection_wb, "ZeroCollection")
    save_wb(CorpCollection_wb, "CorpCollection")
    save_wb(NonCorpCollection_wb, "NonCorpCollection")
    save_wb(AMERCollection_wb, "AMERCollection")
    save_wb(MexicoCollection_wb, "MexicoCollection")
    save_wb(APACCollection_wb, "APACCollection")
    save_wb(EMEAACollection_wb, "EMEAACollection")
    save_wb(GCCollection_wb, "GCCollection")
    save_wb(AMER_CROP_wb, "AMER_CROP")
    save_wb(AMER_NONCROP_wb, "AMER_NONCROP")
    save_wb(MEXICO_CROP_wb, "MEXICO_CROP")
    save_wb(MEXICO_NONCROP_wb, "MEXICO_NONCROP")
    save_wb(APAC_CROP_wb, "APAC_CROP")
    save_wb(APAC_NONCROP_wb, "APAC_NONCROP")
    save_wb(EMEAA_CROP_wb, "EMEAA_CROP")
    save_wb(EMEAA_NONCROP_wb, "EMEAA_NONCROP")
    save_wb(GC_CROP_wb, "GC_CROP")
    save_wb(GC_NONCROP_wb, "GC_NONCROP")
    save_wb(APAC_AMEA_COMBINED_wb, "APAC_AMEA_COMBINED")
    save_wb(APAC_AMEA_CROP_wb, "APAC_AMEA_CROP")
    save_wb(APAC_AMEA_NONCROP_wb, "APAC_AMEA_NONCROP")
    save_wb(APAC_AMEA_GC_COMBINED_wb, "APAC_AMEA_GC_COMBINED")
    save_wb(APAC_AMEA_GC_CROP_wb, "APAC_AMEA_GC_CROP")
    save_wb(APAC_AMEA_GC_NONCROP_wb, "APAC_AMEA_GC_NONCROP")


def _split_region_collections(cleaned_workbook: Workbook, source_stem: str) -> dict[str, str]:
    output_dir = Path(settings.output_dir) / "Region_Wise_Split"
    output_dir.mkdir(parents=True, exist_ok=True)

    # Only process the first sheet for simplicity (can be extended for multi-sheet)
    cleaned_sheet = cleaned_workbook.worksheets[0]
    rows = list(
        cleaned_sheet.iter_rows(
            min_row=1,
            max_row=cleaned_sheet.max_row,
            max_col=cleaned_sheet.max_column,
            values_only=True,
        )
    )
    if not rows:
        return {}
    header = list(rows[0])
    col_idx = {str(h).strip().upper(): i for i, h in enumerate(header)}

    # Create output workbooks and sheets
    def new_ws(name):
        wb = Workbook()
        ws = wb.active
        ws.title = name
        ws.append(header)
        return wb, ws
    AMER_wb, AMER_ws = new_ws("AMER")
    AMER_wb_input, AMER_ws_input = new_ws("AMER_INPUT")  # For AMER PeopleSoft input if needed
    APAC_wb, APAC_ws = new_ws("APAC")
    GC_wb, GC_ws = new_ws("GC")
    APAC_GC_wb, APAC_GC_ws = new_ws("APAC_GC")
    EMEAA_wb, EMEAA_ws = new_ws("EMEAA")

    # Loop through rows and distribute as per VBO logic
    for row in rows[1:]:
        bu = str(row[col_idx.get("BU", -1)]).strip().upper() if col_idx.get("BU") is not None else ""
        region = str(row[col_idx.get("REGION", -1)]).strip().upper() if col_idx.get("REGION") is not None else ""
        if bu.startswith("A") :
            AMER_ws.append(row)
        elif bu.startswith("A") and (region == "GC" or region == "APAC" or region == "EMEAA" or region == "AMEA"):
            AMER_ws_input.append(row)
        elif bu.startswith("P"):
            if region == "APAC" or region == "GC" or region == "AMEA":
                APAC_GC_ws.append(row)
            elif region == "GC":
                GC_ws.append(row)
            elif region == "APAC" or region == "AMEA":
                APAC_ws.append(row)
            else:
                APAC_ws.append(row)
        elif bu.startswith("H"):
            EMEAA_ws.append(row)

    # Save all workbooks
    output_paths: dict[str, str] = {}
    def save_wb(wb, name):
        path = _next_available_path(output_dir / f"{name}_{source_stem}.xlsx")
        wb.save(path)
        output_paths[name] = str(path)
        logger.info("Created %s region file: %s", name, path)
    save_wb(AMER_wb, "AMER")
    save_wb(AMER_wb_input, "AMER_INPUT")
    save_wb(APAC_wb, "APAC")
    save_wb(GC_wb, "GC")
    save_wb(APAC_GC_wb, "APAC_GC")
    save_wb(EMEAA_wb, "EMEAA")
    return output_paths


def _split_intercompany_collections(cleaned_workbook: Workbook, source_stem: str) -> None:
    output_dir = Path(settings.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    intercompany_workbook = Workbook()
    normal_workbook = Workbook()
    intercompany_workbook.remove(intercompany_workbook.active)
    normal_workbook.remove(normal_workbook.active)

    for cleaned_sheet in cleaned_workbook.worksheets:
        intercompany_sheet = intercompany_workbook.create_sheet(title=cleaned_sheet.title)
        normal_sheet = normal_workbook.create_sheet(title=cleaned_sheet.title)

        rows = list(
            cleaned_sheet.iter_rows(
                min_row=1,
                max_row=cleaned_sheet.max_row,
                max_col=cleaned_sheet.max_column,
                values_only=True,
            )
        )

        if not rows:
            continue

        header = list(rows[0])
        intercompany_sheet.append(header)
        normal_sheet.append(header)

        bu_index = _find_column(header, "BU")

        for row in rows[1:]:
            row_values = list(row)
            bu = ""

            if bu_index is not None and bu_index < len(row_values):
                bu = str(row_values[bu_index]).strip().upper()

            if bu.startswith("H") or bu.startswith("A"):
                intercompany_sheet.append(row_values)
            else:
                normal_sheet.append(row_values)

    intercompany_path = _next_available_path(output_dir / f"Intercompany_{source_stem}.xlsx")
    normal_path = _next_available_path(output_dir / f"Normal_{source_stem}.xlsx")
    intercompany_workbook.save(intercompany_path)
    normal_workbook.save(normal_path)

    logger.info("Created Intercompany file: %s", intercompany_path)
    logger.info("Created Normal file: %s", normal_path)


def _next_available_path(path: Path) -> Path:
    if not path.exists():
        return path
    stem = path.stem
    suffix = path.suffix
    index = 1
    while True:
        candidate = path.parent / f"{stem}_{index}{suffix}"
        if not candidate.exists():
            return candidate
        index += 1


def _is_red_fill(cell) -> bool:
    fill = getattr(cell, "fill", None)
    if not fill or fill.fill_type is None:
        return False

    color = fill.start_color
    if not color:
        return False

    # RGB colors (most common in modern .xlsx files)
    if color.type == "rgb" and color.rgb:
        rgb = color.rgb.upper()
        return rgb in {"FFFF0000", "FF0000", "00FF0000"}

    # Indexed palette red (legacy formats/styles)
    if color.type == "indexed" and color.indexed is not None:
        return int(color.indexed) == 10

    return False


def remove_red_rows_from_excel(
    input_file_path: str,
    output_dir: str | None = None,
) -> str:
    """Create a new Excel file excluding rows with red-filled cells."""
    _ensure_expected_output_dirs()

    source_path = Path(input_file_path)
    target_dir = Path(output_dir or settings.upload_dir)
    target_dir.mkdir(parents=True, exist_ok=True)

    input_workbook = load_workbook(source_path)
    output_workbook = Workbook()
    output_workbook.remove(output_workbook.active)

    for input_sheet in input_workbook.worksheets:
        output_sheet = output_workbook.create_sheet(title=input_sheet.title)

        for row in input_sheet.iter_rows(
            min_row=1,
            max_row=input_sheet.max_row,
            max_col=input_sheet.max_column,
        ):
            # Keep header row as-is.
            if row[0].row == 1:
                output_sheet.append([cell.value for cell in row])
                continue

            if not any(_is_red_fill(cell) for cell in row):
                output_sheet.append([cell.value for cell in row])


    output_path = target_dir / f"cleaned_no_red_{source_path.stem}.xlsx"
    output_path = _next_available_path(output_path)
    output_workbook.save(output_path)

    # === NEW: Copy data from Manual_entry file(s) if any exist ===
    manual_entry_dir = Path(settings.upload_dir) / "Manual_entry" if hasattr(settings, "upload_dir") else Path("data/Manual_entry")
    manual_files = list(manual_entry_dir.glob("*.xlsx"))
    if manual_files:
        from openpyxl import load_workbook as _load_wb
        # Use the first file found (can be extended to merge all if needed)
        manual_file = manual_files[0]
        manual_wb = _load_wb(manual_file)
        manual_ws = manual_wb.active
        # Open the cleaned file again for appending
        cleaned_wb = load_workbook(output_path)
        cleaned_ws = cleaned_wb.active
        # Find the header row in manual_ws (assume first row)
        manual_header = [cell.value for cell in next(manual_ws.iter_rows(min_row=1, max_row=1))]
        cleaned_header = [cell.value for cell in next(cleaned_ws.iter_rows(min_row=1, max_row=1))]
        # Map columns by header name
        manual_col_map = {str(h).strip().upper(): i for i, h in enumerate(manual_header)}
        cleaned_col_map = {str(h).strip().upper(): i for i, h in enumerate(cleaned_header)}
        # For each data row in manual_ws, append to cleaned_ws in correct order
        for row in manual_ws.iter_rows(min_row=2, max_row=manual_ws.max_row, values_only=True):
            # Build row for cleaned_ws
            new_row = [None] * len(cleaned_header)
            for h, idx in cleaned_col_map.items():
                if h in manual_col_map:
                    val = row[manual_col_map[h]]
                    new_row[idx] = val
            cleaned_ws.append(new_row)
        cleaned_wb.save(output_path)
        logger.info(f"Appended data from Manual_entry file {manual_file} to cleaned file {output_path}")

    # Continue with further processing
    corp_noncorp_split = _split_user_type_collections(cleaned_workbook=load_workbook(output_path), source_stem=source_path.stem)
    region_output_paths = _split_region_collections(cleaned_workbook=load_workbook(output_path), source_stem=source_path.stem)
    _split_intercompany_collections(cleaned_workbook=load_workbook(output_path), source_stem=source_path.stem)

    try:
        generate_amer_peoplesoft_output(input_file_path=str(output_path))
    except Exception as exc:
        logger.warning("Failed AMER PeopleSoft generation for cleaned file %s: %s", output_path, exc)

    try:
        amer_file = region_output_paths.get("AMER_INPUT")
        if amer_file:
            generate_amer_intercompany_output(input_file_path=amer_file)
        else:
            logger.warning("AMER region file not found in region split output")
    except Exception as exc:
        logger.warning("Failed AMER Intercompany generation for cleaned file %s: %s", output_path, exc)

    apac_processing_result: dict[str, str | int] | None = None
    try:
        apac_gc_file = region_output_paths.get("APAC_GC")
        if apac_gc_file:
            apac_processing_result = generate_apac_processing_output(input_file_path=apac_gc_file)
        else:
            logger.warning("APAC GC not found into into region wise split folder path")    
        # apac_processing_result = generate_apac_processing_output(input_file_path=str(output_path))
        
    except Exception as exc:
        logger.warning("Failed APAC processing generation for cleaned file %s: %s", output_path, exc)

    input_apac_gc_intewrcompany_result: dict[str, str | int] | None = None 
    try:
        input_apac_gc_intewrcompany_result = generate_input_apac_gc_intewrcompany(input_file_path=str(output_path))
    except Exception as exc:
        logger.warning("Failed APAC processing generation for cleaned file %s: %s", output_path, exc)    

    try:
        if input_apac_gc_intewrcompany_result:
            generate_apac_gc_intewrcompany_output(input_file_path=input_apac_gc_intewrcompany_result)
        else:
            logger.warning("APAC GC INTERCOMPANY input not found from input service")
    except Exception as exc:
        logger.warning("Failed APAC GC Intercompany generation for cleaned file %s: %s", output_path, exc)

    try:
        gaf_noncrop_file = (
            str(apac_processing_result["gaf_noncrop_path"])
            if apac_processing_result and apac_processing_result.get("gaf_noncrop_path")
            else None
        )
        if gaf_noncrop_file:
            generate_gaf_apac_output(input_file_path=gaf_noncrop_file)
        else:
            logger.warning("GAF NONCROP collection not found in APAC processing output")
    except Exception as exc:
        logger.warning("Failed GAF APAC generation for cleaned file %s: %s", output_path, exc)

    try:
        rir_noncrop_file = (
            str(apac_processing_result["rir_noncrop_path"])
            if apac_processing_result and apac_processing_result.get("rir_noncrop_path")
            else None
        )
        if rir_noncrop_file:
            generate_rir_apac_output(input_file_path=rir_noncrop_file)
        else:
            logger.warning("RIR GC NONCROP collection not found in APAC processing output")
    except Exception as exc:
        logger.warning("Failed APAC_GC_RIR generation for cleaned file %s: %s", output_path, exc)

        

    emeaa_processing_result: dict[str, str | int] | None = None
    try:
        emeaa_file = region_output_paths.get("EMEAA")
        if emeaa_file:
            emeaa_processing_result = generate_emeaa_processing_output(input_file_path=emeaa_file)
        else:
            logger.warning("EMEAA region file not found in region split output")
    except Exception as exc:
        logger.warning("Failed EMEAA processing generation for cleaned file %s: %s", output_path, exc)

    input_emeaa_intercompany_result: dict[str, str | int] | None = None
    try:
        input_emeaa_intercompany_result = generate_input_emeaa_intercompany_result(input_file_path=output_path)
    except Exception as exc:
        logger.warning("Failed EMEAA processing generation for cleaned file %s: %s", output_path, exc)

    try:
        if input_emeaa_intercompany_result:
            generate_emeaa_intercompany_output(input_file_path=input_emeaa_intercompany_result)
        else:
            logger.warning("EMEAA_INTERCOMPANY collection not found in EMEAA processing output")
    except Exception as exc:
        logger.warning("Failed EMEAA Intercompany generation for cleaned file %s: %s", output_path, exc)


    # Only generate JRF output if there is at least one data row (excluding header)
    try:
        gc_corp_billing = corp_noncorp_split.get("GC_CROP")
        if gc_corp_billing:
            generate_jrf_output(input_file_path=gc_corp_billing)
        else:
            logger.warning("RIR GC CROP collection not found in APAC processing output")
    except Exception as exc:
        logger.warning("Failed APAC_GC_RIR generation for cleaned file %s: %s", output_path, exc)


    # try:
    #     has_data = False
    #     for sheet in output_workbook.worksheets:
    #         rows = list(sheet.iter_rows(min_row=2, max_row=sheet.max_row, max_col=sheet.max_column, values_only=True))
    #         # Check if any row has at least one non-None value
    #         if any(any(cell is not None for cell in row) for row in rows):
    #             has_data = True
    #             break
    #     if has_data:
    #         generate_jrf_output(input_file_path=str(output_path))
    #     else:
    #         logger.info("No data found for JRF output in cleaned file %s; skipping JRF output generation.", output_path)
    # except Exception as exc:
    #     logger.warning("Failed JRF generation for cleaned file %s: %s", output_path, exc)

    logger.info("Created cleaned Excel file without red rows: %s", output_path)
    return str(output_path)