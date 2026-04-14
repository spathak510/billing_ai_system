from __future__ import annotations

import logging
from datetime import datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
import calendar

import pandas as pd
from openpyxl import load_workbook

from app.config.settings import settings

logger = logging.getLogger(__name__)

_RIR_REQUIRED_SHEETS = ("upload sheet", "Recharge Form")


def _resolve_input_path(input_file_path: str | None) -> Path:
    """Resolve the input file path for RIR APAC processing."""
    if input_file_path:
        return Path(input_file_path)

    collection_dir = Path(settings.output_dir) / "APAC" / "APAC_Output"
    collection_files = sorted(
        collection_dir.glob("*_RIR_NONCROP.xlsx"),
        key=lambda p: p.stat().st_mtime,
        reverse=True,
    )
    if collection_files:
        return collection_files[0]

    legacy_cleaned_files = sorted(
        Path(settings.output_dir).glob("cleaned_no_red_*.xlsx"),
        key=lambda p: p.stat().st_mtime,
        reverse=True,
    )
    if legacy_cleaned_files:
        return legacy_cleaned_files[0]

    raise FileNotFoundError("No *_RIR_NONCROP.xlsx file found in output/APAC/APAC_Output.")


def _find_col(columns: list[str], candidates: list[str]) -> str | None:
    """Find a column by trying multiple candidate names (case-insensitive)."""
    normalized = {str(col).strip().upper().replace(" ", ""): col for col in columns}
    for candidate in candidates:
        key = candidate.upper().replace(" ", "")
        if key in normalized:
            return normalized[key]
    return None


def _to_decimal(value: object) -> Decimal:
    """Convert a value to Decimal, handling commas and empty strings."""
    text = str(value).strip().replace(",", "")
    if not text:
        return Decimal("0")
    try:
        return Decimal(text)
    except InvalidOperation:
        return Decimal("0")


def _set_cell_value_safe(ws, cell_ref: str, value: object) -> None:
    """Set cell value safely, handling merged cells."""
    cell = ws[cell_ref]
    if cell.__class__.__name__ != "MergedCell":
        cell.value = value
        return

    for merged_range in ws.merged_cells.ranges:
        if cell_ref in merged_range:
            ws.cell(row=merged_range.min_row, column=merged_range.min_col).value = value
            return

    ws[cell_ref].value = value


def _resolve_template_path(template_path: str | None) -> Path:
    if template_path:
        resolved = Path(template_path)
        if not resolved.exists():
            raise FileNotFoundError(f"Template file not found: {resolved}")
        return resolved

    template_dir = Path(settings.output_dir) / "APAC" / "APAC_GC_RIR" / "Template_Formate"
    preferred_template = template_dir / "RIR_GC_APAC_NON-CORP_FEBRUARY26.xlsx"
    if preferred_template.exists():
        return preferred_template

    template_files = sorted(template_dir.glob("*.xlsx"), key=lambda p: p.stat().st_mtime, reverse=True)
    if template_files:
        return template_files[0]

    legacy_dir = Path(settings.output_dir) / "APAC" / "APAC_GC_RIR"
    legacy_templates = sorted(legacy_dir.glob("*.xlsx"), key=lambda p: p.stat().st_mtime, reverse=True)
    if legacy_templates:
        return legacy_templates[0]

    raise FileNotFoundError("RIR APAC template not found in output/APAC/APAC_GC_RIR/Template_Formate.")


def generate_rir_apac_output(
    input_file_path: str | None = None,
    template_path: str | None = None,
    output_folder_path: str | None = None,
    submitted_by: str = "",
    output_file_name: str = "RIR_GC_APAC_NON-CORP",
) -> dict[str, str | int | float]:
    """Generate RIR APAC workbook from cleaned data using the provided template format.
    
    Args:
        input_file_path: Path to the cleaned data file (optional, auto-resolves if None)
        submitted_by: Name of the user submitting the data
        output_file_name: Base name for the output file
    
    Returns:
        Dictionary with output_path, record count, and total amount
    """
    submitted_by_value = (
        submitted_by.strip() if isinstance(submitted_by, str) and submitted_by.strip() else "GenWizard_Automation"
    )

    source_path = _resolve_input_path(input_file_path)
    df = pd.read_excel(source_path, sheet_name=0)

    cols = list(df.columns)
    
    # Find required columns with multiple candidate names
    bu_col = _find_col(cols, ["BU"])
    currency_col = _find_col(cols, ["CURRENCYCODE", "CURRENCY CODE", "CURRENCY"])
    amount_col = _find_col(cols, ["AMOUNT", "BILL_AMOUNT", "BILLING_AMOUNT", "TOTAL_AMOUNT", "NET_AMOUNT", "VALUE"])
    holidex_col = _find_col(cols, ["HOLIDEX", "CUSTID", "CUSTID #"])
    course_col = _find_col(cols, ["COURSE_NAME", "COURSE NAME", "DESCRIPTION"])
    order_col = _find_col(cols, ["ORDER_NO", "ORDER NO", "ORDER_NUMBER", "ORDER NUMBER"])
    offering_date_col = _find_col(cols, ["OFFERING_DATE", "OFFERING DATE"])
    employee_col = _find_col(cols, ["EMPLOYEE", "LEARNERS NAME", "LEARNER", "USERNAME"])

    # Check for required columns
    if not bu_col or not currency_col or not amount_col:
        missing = [
            key
            for key, value in {
                "BU": bu_col,
                "CURRENCYCODE": currency_col,
                "AMOUNT": amount_col,
            }.items()
            if value is None
        ]
        raise ValueError(f"Missing required columns for RIR APAC processing: {', '.join(missing)}")

    # Find APPLY REVENUE column
    apply_revenue_col = None
    for col in cols:
        if "APPLYREVENUE" in str(col).upper().replace(" ", ""):
            apply_revenue_col = col
            break

    # Filter rows for RIR APAC processing
    rir_rows: list[dict[str, object]] = []

    for _, row in df.iterrows():
        bu = str(row.get(bu_col, "")).strip().upper()
        amount = _to_decimal(row.get(amount_col, "0"))

        # Skip zero amounts and certain BU types
        if amount == 0:
            continue
        if bu.startswith("H") or bu.startswith("A"):
            continue
        if not bu.startswith("P"):
            continue

        # Add to RIR rows
        row_dict = row.to_dict()
        row_dict[amount_col] = float(amount)
        rir_rows.append(row_dict)

    # Find or create output directory
    output_root = (
        Path(output_folder_path)
        if output_folder_path
        else Path(settings.output_dir) / "APAC" / "APAC_GC_RIR" / "Output"
    )
    output_root.mkdir(parents=True, exist_ok=True)

    # Load template workbook from APAC_GC_RIR template folder.
    resolved_template_path = _resolve_template_path(template_path)
    
    logger.info("Using RIR APAC template: %s", resolved_template_path)

    workbook = load_workbook(resolved_template_path)
    
    # Verify required sheets exist
    if "upload sheet" not in workbook.sheetnames:
        raise ValueError("Template must contain 'upload sheet' sheet")
    if "Recharge Form" not in workbook.sheetnames:
        raise ValueError("Template must contain 'Recharge Form' sheet")

    extra_sheets = [sheet for sheet in workbook.sheetnames if sheet not in _RIR_REQUIRED_SHEETS]
    if extra_sheets:
        logger.info(
            "Ignoring extra sheets in RIR workbook %s: %s",
            resolved_template_path,
            ", ".join(extra_sheets),
        )

    upload_sheet = workbook["upload sheet"]
    recharge_sheet = workbook["Recharge Form"]

    # Clear old data rows (rows 10-10000)
    for row_idx in range(10, 10001):
        for col_idx in range(1, 30):
            cell = upload_sheet.cell(row=row_idx, column=col_idx)
            cell.value = None

    # Set header information
    current_dt = datetime.now()
    date_str = f"{current_dt.month}/{current_dt.day}/{current_dt.year}"
    
    _set_cell_value_safe(recharge_sheet, "F8", date_str)
    _set_cell_value_safe(recharge_sheet, "O8", submitted_by_value)

    # Process data rows
    total = Decimal("0")
    row_index = 10
    line_no = 1
    record_count = 0

    for row in rir_rows:
        amount = _to_decimal(row.get(amount_col, 0))

        # Write row data
        upload_sheet.cell(row=row_index, column=1).value = line_no
        upload_sheet.cell(row=row_index, column=2).value = row.get(bu_col, "")
        upload_sheet.cell(row=row_index, column=3).value = row.get(holidex_col, "") if holidex_col else ""
        upload_sheet.cell(row=row_index, column=5).value = date_str
        upload_sheet.cell(row=row_index, column=6).value = "LMS Training"
        upload_sheet.cell(row=row_index, column=7).value = row.get(course_col, "") if course_col else ""
        upload_sheet.cell(row=row_index, column=8).value = row.get(currency_col, "")
        upload_sheet.cell(row=row_index, column=9).value = float(amount)
        upload_sheet.cell(row=row_index, column=11).value = row.get(apply_revenue_col, "") if apply_revenue_col else ""
        
        # Column 15 (O): COURSE_NAME
        upload_sheet.cell(row=row_index, column=15).value = row.get(course_col, "") if course_col else ""
        
        # Column 26: ORDER_NO
        upload_sheet.cell(row=row_index, column=26).value = row.get(order_col, "") if order_col else ""
        
        # Column 28 (AB): OFFERING_DATE
        upload_sheet.cell(row=row_index, column=28).value = row.get(offering_date_col, "") if offering_date_col else ""
        
        # Column 29 (AC): EMPLOYEE
        upload_sheet.cell(row=row_index, column=29).value = row.get(employee_col, "") if employee_col else ""

        total += amount
        record_count += 1
        row_index += 1
        line_no += 1

    # Set totals in upload sheet (I5)
    _set_cell_value_safe(upload_sheet, "I5", float(total))

    # Set totals in recharge form (I33)
    _set_cell_value_safe(recharge_sheet, "I33", float(total))

    # Save with fixed output name (overwrite existing file), matching VBO behavior.
    base_name = output_file_name.strip() or "RIR_GC_APAC_NON-CORP"
    if base_name.lower().endswith(".xlsx"):
        base_name = base_name[:-5]

    final_file_name = f"{base_name}_{current_dt.strftime('%B %Y')}.xlsx"
    output_path = output_root / final_file_name

    if output_path.exists():
        output_path.unlink()

    workbook.save(output_path)

    logger.info("Generated RIR APAC output: %s (records: %d, total: %.2f)", output_path, record_count, float(total))
    return {
        "rir_apac_output_path": str(output_path),
        "rir_apac_records": record_count,
        "rir_apac_total": float(total),
        "template_file": str(resolved_template_path),
        "source_file": str(source_path),
    }
