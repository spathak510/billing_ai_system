from __future__ import annotations

import logging
from decimal import Decimal, InvalidOperation
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from app.config.settings import settings

logger = logging.getLogger(__name__)


def _normalized_key(value: object) -> str:
    return str(value).strip().upper().replace(" ", "")


def _set_cell_value_safe(ws, cell_ref: str, value: object) -> None:
    cell = ws[cell_ref]
    if cell.__class__.__name__ != "MergedCell":
        cell.value = value
        return

    for merged_range in ws.merged_cells.ranges:
        if cell_ref in merged_range:
            ws.cell(row=merged_range.min_row, column=merged_range.min_col).value = value
            return

    ws[cell_ref].value = value


def _resolve_input_path(input_file_path: str | None) -> Path:
    if input_file_path:
        return Path(input_file_path)

    output_dir = Path(settings.output_dir)
    cleaned_files = sorted(output_dir.glob("cleaned_no_red_*.xlsx"), key=lambda p: p.stat().st_mtime, reverse=True)
    if cleaned_files:
        return cleaned_files[0]

    raise FileNotFoundError("No cleaned_no_red_*.xlsx file found in output folder.")


def _resolve_apac_gc_intercompany_input_path(input_file_path: str | None) -> Path:
    if input_file_path:
        return Path(input_file_path)

    collection_dir = Path(settings.output_dir) / "APAC" / "APAC_Output"
    collection_files = sorted(
        collection_dir.glob("*_APAC_GC_NONCROP.xlsx"),
        key=lambda p: p.stat().st_mtime,
        reverse=True,
    )
    if collection_files:
        return collection_files[0]

    raise FileNotFoundError("No *_APAC_GC_NONCROP.xlsx file found in APAC_Output.")


def _resolve_apac_gc_intercompany_template_path(template_path: str | None) -> Path:
    if template_path:
        resolved = Path(template_path)
        if not resolved.exists():
            raise FileNotFoundError(f"Template file not found: {resolved}")
        return resolved

    output_root = Path(settings.output_dir)
    template_dirs = [
        output_root / "APAC" / "APAC_Intercompny" / "Template_Format",
        output_root / "APAC" / "APAC_Intercompny" / "Template_Formate",
        output_root / "APAC" / "APAC_Intercompny" / "Template_formate",
    ]
    for template_dir in template_dirs:
        apac_template_files = sorted(
            template_dir.glob("*.xlsx"),
            key=lambda p: p.stat().st_mtime,
            reverse=True,
        )
        if apac_template_files:
            return apac_template_files[0]

    legacy_template_dirs = [
        output_root / "APAC" / "APAC_Output" / "APAC_Intercompny" / "Template_Format",
        output_root / "APAC" / "APAC_Output" / "APAC_Intercompny" / "Template_Formate",
        output_root / "APAC" / "APAC_Output" / "APAC_Intercompny" / "Template_formate",
    ]
    for legacy_apac_template_dir in legacy_template_dirs:
        legacy_apac_template_files = sorted(
            legacy_apac_template_dir.glob("*.xlsx"),
            key=lambda p: p.stat().st_mtime,
            reverse=True,
        )
        if legacy_apac_template_files:
            return legacy_apac_template_files[0]

    amer_template_dir = Path(settings.output_dir) / "AMER_Intercompny" / "Template_Format"
    template_files = sorted(
        amer_template_dir.glob("*.xlsx"),
        key=lambda p: p.stat().st_mtime,
        reverse=True,
    )
    if template_files:
        return template_files[0]

    amer_dir = Path(settings.output_dir) / "AMER_Intercompny"
    legacy_template_files = sorted(
        amer_dir.glob("*.xlsx"),
        key=lambda p: p.stat().st_mtime,
        reverse=True,
    )
    if legacy_template_files:
        return legacy_template_files[0]

    raise FileNotFoundError(
        "No intercompany template workbook found in APAC/AMER Template_Format folders."
    )


def _to_decimal(value: object) -> Decimal:
    text = str(value).strip().replace(",", "")
    if not text:
        return Decimal("0")
    try:
        return Decimal(text)
    except InvalidOperation:
        return Decimal("0")


def _output_stem(source_stem: str, custom_stem: str | None) -> str:
    if custom_stem and custom_stem.strip():
        return custom_stem.strip()
    if source_stem.upper().startswith("CLEANED_NO_RED_"):
        return "APAC Processing"
    return source_stem


def _value_from_row(row_dict: dict[str, object], column_name: str) -> object:
    normalized_target = _normalized_key(column_name)
    for key, value in row_dict.items():
        if _normalized_key(key) == normalized_target:
            return value
    return ""


def _clear_billing_lines(ws) -> None:
    if ws.max_row <= 1:
        return

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=26):
        for cell in row:
            cell.value = None


def _find_revenue_column(columns: list[object]) -> str | None:
    for col in columns:
        if "REVEN" in _normalized_key(col):
            return str(col)
    return None


def generate_apac_processing_output(
    input_file_path: str | None = None,
    output_stem: str | None = None,
) -> dict[str, str | int]:
    """Process APAC rules from cleaned data and write all APAC collection outputs."""
    source_path = _resolve_input_path(input_file_path)
    df = pd.read_excel(source_path, sheet_name=0)

    col_map = {str(col).upper(): col for col in df.columns}
    bu_col = col_map.get("BU")
    currency_col = col_map.get("CURRENCYCODE")
    user_type_col = col_map.get("USER_TYPE")
    amount_col = col_map.get("AMOUNT")

    if not bu_col or not currency_col or not user_type_col or not amount_col:
        missing = [
            key
            for key, col in {
                "BU": bu_col,
                "CURRENCYCODE": currency_col,
                "USER_TYPE": user_type_col,
                "AMOUNT": amount_col,
            }.items()
            if col is None
        ]
        raise ValueError(f"Missing required columns for APAC processing: {', '.join(missing)}")

    apac_gc_crop_rows: list[dict[str, object]] = []
    apac_gc_noncrop_rows: list[dict[str, object]] = []
    rir_apac_crop_rows: list[dict[str, object]] = []
    rir_gc_crop_rows: list[dict[str, object]] = []
    rir_noncrop_rows: list[dict[str, object]] = []
    gaf_noncrop_rows: list[dict[str, object]] = []

    for _, row in df.iterrows():
        bu = str(row.get(bu_col, "")).strip().upper()
        currency = str(row.get(currency_col, "")).strip().upper()
        user_type = str(row.get(user_type_col, "")).strip().upper()
        amount = _to_decimal(row.get(amount_col, "0"))

        row_dict = row.to_dict()

        if not bu.startswith("P"):
            continue

        if currency == "EUR":
            amount = (amount / Decimal("0.86")).quantize(Decimal("0.01"))
            currency = "USD"

        row_dict[amount_col] = float(amount)
        row_dict[currency_col] = currency

        if currency not in {"USD", "CNY"}:
            continue

        if user_type == "C":
            apac_gc_crop_rows.append(row_dict)
        else:
            apac_gc_noncrop_rows.append(row_dict)

    for row_dict in apac_gc_crop_rows:
        amount = _to_decimal(row_dict.get(amount_col, "0"))
        if amount > 0:
            rir_apac_crop_rows.append(row_dict)

    for row_dict in apac_gc_noncrop_rows:
        amount = _to_decimal(row_dict.get(amount_col, "0"))
        if amount > 0:
            rir_noncrop_rows.append(row_dict)
        elif amount < 0:
            gaf_noncrop_rows.append(row_dict)

    output_dir = Path(settings.output_dir) / "APAC" / "APAC_Output"
    output_dir.mkdir(parents=True, exist_ok=True)
    stem = _output_stem(source_path.stem, output_stem)

    collections: dict[str, list[dict[str, object]]] = {
        "APAC_GC_CROP": apac_gc_crop_rows,
        "APAC_GC_NONCROP": apac_gc_noncrop_rows,
        "RIR_APAC_CROP": rir_apac_crop_rows,
        "RIR_GC_CROP": rir_gc_crop_rows,
        "RIR_NONCROP": rir_noncrop_rows,
        "GAF_NONCROP": gaf_noncrop_rows,
    }

    result: dict[str, str | int] = {
        "apac_gc_crop_rows": len(apac_gc_crop_rows),
        "apac_gc_noncrop_rows": len(apac_gc_noncrop_rows),
        "rir_apac_crop_rows": len(rir_apac_crop_rows),
        "rir_gc_crop_rows": len(rir_gc_crop_rows),
        "rir_noncrop_rows": len(rir_noncrop_rows),
        "gaf_noncrop_rows": len(gaf_noncrop_rows),
    }

    for name, rows in collections.items():
        collection_path = output_dir / f"{stem}_{name}.xlsx"
        pd.DataFrame(rows, columns=df.columns).to_excel(collection_path, index=False)
        result[f"{name.lower()}_path"] = str(collection_path)

    logger.info("Generated APAC processing outputs in %s", output_dir)
    return result







def generate_apac_gc_intewrcompany_output(
    input_file_path: str | None = None,
    template_path: str | None = None,
    output_folder_path: str | None = None,
    base_file_name: str | None = None,
    request_date: str | None = None,
    request_name: str = "",
    account_number: str = "",
) -> dict[str, str | int]:
    """Generate APAC GC intercompany workbook from the APAC GC NONCROP collection."""
    source_path = _resolve_apac_gc_intercompany_input_path(input_file_path)
    resolved_template = _resolve_apac_gc_intercompany_template_path(template_path)

    target_dir = (
        Path(output_folder_path)
        if output_folder_path
        else Path(settings.output_dir) / "APAC" / "APAC_Intercompny" / "Output"
    )
    target_dir.mkdir(parents=True, exist_ok=True)

    output_base_name = (base_file_name or "APAC_GC_Intercompany billing lines").strip()
    if output_base_name.lower().endswith(".xlsx"):
        output_base_name = output_base_name[:-5]
    final_file_name = f"{output_base_name}_{datetime.now().strftime('%B %Y')}.xlsx"
    output_path = target_dir / final_file_name

    df = pd.read_excel(source_path, sheet_name=0)
    if df.empty:
        raise ValueError("BillingCollection is EMPTY for APAC GC Intercompany processing.")

    wb = load_workbook(resolved_template)
    if "RIR" not in wb.sheetnames or "BILLING LINES" not in wb.sheetnames:
        raise ValueError(
            f"APAC GC Intercompany template {resolved_template} does not contain RIR and BILLING LINES sheets."
        )

    rir_sheet = wb["RIR"]
    billing_sheet = wb["BILLING LINES"]

    default_request_name = "GenWizard_Automation"
    rir_name_value = request_name.strip() if isinstance(request_name, str) and request_name.strip() else default_request_name

    _set_cell_value_safe(rir_sheet, "F10", request_date or datetime.now().strftime("%Y-%m-%d"))
    _set_cell_value_safe(rir_sheet, "P10", rir_name_value)
    _set_cell_value_safe(rir_sheet, "F11", account_number)

    _clear_billing_lines(billing_sheet)

    rows = df.to_dict(orient="records")
    revenue_col = _find_revenue_column(list(df.columns))

    write_row = 2
    for row_dict in rows:
        billing_sheet.cell(write_row, 1).value = _value_from_row(row_dict, "USERNAME")
        billing_sheet.cell(write_row, 2).value = _value_from_row(row_dict, "EMPLOYEE")
        billing_sheet.cell(write_row, 3).value = _value_from_row(row_dict, "HOLIDEX")
        billing_sheet.cell(write_row, 4).value = _value_from_row(row_dict, "AMOUNT")
        billing_sheet.cell(write_row, 5).value = _value_from_row(row_dict, "CURRENCYCODE")
        billing_sheet.cell(write_row, 6).value = _value_from_row(row_dict, "COST_CENTER")
        billing_sheet.cell(write_row, 7).value = _value_from_row(row_dict, "ORDER_NO")
        billing_sheet.cell(write_row, 8).value = _value_from_row(row_dict, "COURSE_NAME")
        billing_sheet.cell(write_row, 9).value = _value_from_row(row_dict, "FACILITY")
        billing_sheet.cell(write_row, 10).value = _value_from_row(row_dict, "OFFERING_ID")
        billing_sheet.cell(write_row, 11).value = _value_from_row(row_dict, "INSTRUCTOR")
        billing_sheet.cell(write_row, 12).value = _value_from_row(row_dict, "OFFERING_DATE")
        billing_sheet.cell(write_row, 13).value = _value_from_row(row_dict, "COUNTRY")
        billing_sheet.cell(write_row, 14).value = _value_from_row(row_dict, "REGION")
        billing_sheet.cell(write_row, 15).value = _value_from_row(row_dict, "USER_TYPE")
        billing_sheet.cell(write_row, 16).value = _value_from_row(row_dict, "TRANSTYPECODE")
        billing_sheet.cell(write_row, 17).value = _value_from_row(row_dict, "PAY_DATE")
        billing_sheet.cell(write_row, 18).value = _value_from_row(row_dict, "NAME")
        billing_sheet.cell(write_row, 19).value = _value_from_row(row_dict, "DELIVERED_ON")
        billing_sheet.cell(write_row, 20).value = row_dict.get(revenue_col, "") if revenue_col else ""
        billing_sheet.cell(write_row, 21).value = _value_from_row(row_dict, "BU")
        write_row += 1

    if output_path.exists():
        output_path.unlink()

    wb.save(output_path)
    logger.info("Generated APAC GC Intercompany output: %s", output_path)
    return {
        "apac_gc_intercompany_file": str(output_path),
        "apac_gc_intercompany_rows": len(rows),
        "template_file": str(resolved_template),
        "source_file": str(source_path),
    }
