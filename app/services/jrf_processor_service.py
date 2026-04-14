from __future__ import annotations

import logging
from datetime import datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from app.config.settings import settings

logger = logging.getLogger(__name__)


def _resolve_input_path(input_file_path: str | None) -> Path:
    """Resolve the input file path for JRF processing."""
    if input_file_path:
        return Path(input_file_path)

    output_dir = Path(settings.output_dir)
    cleaned_files = sorted(
        output_dir.glob("cleaned_no_red_*.xlsx"), key=lambda p: p.stat().st_mtime, reverse=True
    )
    if cleaned_files:
        return cleaned_files[0]

    raise FileNotFoundError("No cleaned_no_red_*.xlsx file found in output folder.")


def _find_col(columns: list[str], candidates: list[str]) -> str | None:
    """Find a column by trying multiple candidate names (case-insensitive)."""
    normalized = {str(col).strip().upper().replace(" ", ""): col for col in columns}
    for candidate in candidates:
        key = candidate.upper().replace(" ", "")
        if key in normalized:
            return normalized[key]
    return None


def _to_decimal(value: object) -> Decimal:
    """Convert a value to Decimal, handling commas and empty strings."""
    text = str(value).strip().replace(",", "")
    if not text:
        return Decimal("0")
    try:
        return Decimal(text)
    except InvalidOperation:
        return Decimal("0")


def _set_cell_value_safe(ws, cell_ref: str, value: object) -> None:
    cell = ws[cell_ref]
    if cell.__class__.__name__ != "MergedCell":
        cell.value = value
        return

    for merged_range in ws.merged_cells.ranges:
        if cell_ref in merged_range:
            ws.cell(row=merged_range.min_row, column=merged_range.min_col).value = value
            return

    ws[cell_ref].value = value


def _resolve_template_path() -> Path:
    output_root = Path(settings.output_dir) / "JRF"
    template_dirs = [
        output_root / "Template_Formate",
        output_root / "Template_Format",
        output_root,
    ]

    for template_dir in template_dirs:
        template_candidates = sorted(template_dir.glob("*.xlsm"), key=lambda p: p.stat().st_mtime, reverse=True)
        if template_candidates:
            return template_candidates[0]

    raise FileNotFoundError(
        f"JRF template (.xlsm) not found in {output_root / 'Template_Formate'} or {output_root}. "
        f"Please place a template file in output/JRF/Template_Formate."
    )


def generate_jrf_output(
    input_file_path: str | None = None,
    requester_name: str = "",
    sheet_name: str = "Journal Template",
) -> dict[str, str | int]:
    """Generate JRF (Journal Reference File) workbook from cleaned data using template format.
    
    The JRF processor creates journal entries in two steps:
    1. All negative entries (with APPLY REVENUE account)
    2. All positive entries (with COST_CENTER account, fallback to APPLY REVENUE if empty)
    
    Args:
        input_file_path: Path to the cleaned data file (optional, auto-resolves if None)
        requester_name: Name of the user requesting the JRF
        sheet_name: Target sheet name in the template (default: "Sheet1")
    
    Returns:
        Dictionary with output_path and row count
    """
    source_path = _resolve_input_path(input_file_path)
    df = pd.read_excel(source_path, sheet_name=0)

    cols = list(df.columns)
    
    # Find required columns with multiple candidate names
    bu_col = _find_col(cols, ["BU"])
    amount_col = _find_col(cols, ["AMOUNT", "BILL_AMOUNT", "BILLING_AMOUNT", "TOTAL_AMOUNT", "NET_AMOUNT", "VALUE"])
    apply_revenue_col = _find_col(cols, ["APPLY REVENUE", "APPLYREVENUE", "REVENUE"])
    course_col = _find_col(cols, ["COURSE_NAME", "COURSE NAME", "DESCRIPTION"])
    username_col = _find_col(cols, ["USERNAME", "USER_NAME", "EMPLOYEE", "LEARNERS NAME"])
    currency_col = _find_col(cols, ["CURRENCYCODE", "CURRENCY CODE", "CURRENCY"])
    cost_center_col = _find_col(cols, ["COST_CENTER", "COSTCENTER", "COST CENTER"])

    # Check for required columns
    if not bu_col or not amount_col or not apply_revenue_col:
        missing = [
            key
            for key, value in {
                "BU": bu_col,
                "AMOUNT": amount_col,
                "APPLY REVENUE": apply_revenue_col,
            }.items()
            if value is None
        ]
        raise ValueError(f"Missing required columns for JRF processing: {', '.join(missing)}")

    # Find or create output directory
    output_root = Path(settings.output_dir) / "JRF" / "Output"
    output_root.mkdir(parents=True, exist_ok=True)

    # Load template from Template_Formate (with compatibility fallbacks)
    template_path = _resolve_template_path()
    
    logger.info("Using JRF template: %s", template_path)

    workbook = load_workbook(template_path)
    
    # Verify sheet exists with robust fallbacks for common template sheet names.
    target_sheet = sheet_name
    if target_sheet not in workbook.sheetnames:
        for candidate in ("Journal Template", "JRF", "Sheet1"):
            if candidate in workbook.sheetnames:
                target_sheet = candidate
                logger.warning(
                    "Requested JRF sheet '%s' not found; using '%s' instead.",
                    sheet_name,
                    target_sheet,
                )
                break
        else:
            available_sheets = ", ".join(workbook.sheetnames)
            raise ValueError(
                f"Sheet '{sheet_name}' not found in template. Available sheets: {available_sheets}"
            )

    ws = workbook[target_sheet]

    # Set header information
    current_dt = datetime.now()
    date_str = f"{current_dt.month:02d}/{current_dt.day:02d}/{current_dt.year}"
    
    _set_cell_value_safe(ws, "B5", date_str)
    _set_cell_value_safe(ws, "B11", requester_name)

    # Find and clear old data rows (rows 16 onwards)
    max_row = ws.max_row
    if max_row >= 16:
        for row_idx in range(16, max_row + 1):
            for col_idx in range(1, 13):  # Columns A-L (1-12)
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = None

    # Process data - Step 1: All negative entries (APPLY REVENUE)
    current_row = 16
    total_rows = 0

    for _, row in df.iterrows():
        amount = _to_decimal(row.get(amount_col, "0"))
        apply_revenue = str(row.get(apply_revenue_col, "")).strip()
        
        ws.cell(row=current_row, column=1).value = "Reclass"
        ws.cell(row=current_row, column=2).value = str(row.get(bu_col, "")).strip()
        ws.cell(row=current_row, column=3).value = apply_revenue
        ws.cell(row=current_row, column=4).value = "216015"
        ws.cell(row=current_row, column=7).value = str(row.get(course_col, "") if course_col else "").strip()
        ws.cell(row=current_row, column=8).value = str(row.get(username_col, "") if username_col else "").strip()
        ws.cell(row=current_row, column=9).value = str(row.get(currency_col, "") if currency_col else "").strip()
        ws.cell(row=current_row, column=10).value = float(-amount)  # Negative amount
        ws.cell(row=current_row, column=11).value = 1

        current_row += 1
        total_rows += 1

    # Leave one blank row
    current_row += 1

    # Process data - Step 2: All positive entries (COST_CENTER)
    for _, row in df.iterrows():
        amount = _to_decimal(row.get(amount_col, "0"))
        
        # Get cost center, with fallback to apply revenue
        cost_center = ""
        if cost_center_col:
            cost_center = str(row.get(cost_center_col, "")).strip()
        
        # Fallback to APPLY REVENUE if cost center is empty
        if not cost_center:
            cost_center = str(row.get(apply_revenue_col, "")).strip()
        
        ws.cell(row=current_row, column=1).value = "Reclass"
        ws.cell(row=current_row, column=2).value = str(row.get(bu_col, "")).strip()
        ws.cell(row=current_row, column=3).value = cost_center
        ws.cell(row=current_row, column=4).value = "216015"
        ws.cell(row=current_row, column=7).value = str(row.get(course_col, "") if course_col else "").strip()
        ws.cell(row=current_row, column=8).value = str(row.get(username_col, "") if username_col else "").strip()
        ws.cell(row=current_row, column=9).value = str(row.get(currency_col, "") if currency_col else "").strip()
        ws.cell(row=current_row, column=10).value = float(amount)  # Positive amount
        ws.cell(row=current_row, column=11).value = 1

        current_row += 1
        total_rows += 1

    # Save the workbook with month/year in filename
    month_year = datetime.now().strftime("%B_%Y")
    output_path = output_root / f"Standard_Journal_Template_{month_year}.xlsm"
    
    # Delete existing file if it exists
    if output_path.exists():
        output_path.unlink()
    
    workbook.save(output_path)

    logger.info("Generated JRF output: %s (total entries: %d)", output_path, total_rows)
    return {
        "jrf_output_path": str(output_path),
        "jrf_entries": total_rows,
    }
