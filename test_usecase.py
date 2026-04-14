from __future__ import annotations

import csv
import io
import os
import shutil
import tempfile
import unittest
from urllib.error import HTTPError
from unittest.mock import patch
from datetime import datetime, timedelta, timezone
from pathlib import Path

from openpyxl import load_workbook

from app.config.settings import settings
from app.main import app
from app.services.sharepoint_download_service import SharePointDownloadClient
from app.services.sharepoint_upload_service import SharePointUploadClient


def _sheet_records(xlsx_path: str, sheet_name: str) -> list[dict[str, object]]:
    wb = load_workbook(xlsx_path)
    ws = wb[sheet_name]

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    header = [str(col) if col is not None else "" for col in rows[0]]
    records: list[dict[str, object]] = []
    for row in rows[1:]:
        if all(cell is None for cell in row):
            continue
        records.append({header[idx]: row[idx] for idx in range(len(header))})
    return records


def _unauthorized_http_error(url: str) -> HTTPError:
    return HTTPError(url, 401, "Unauthorized", hdrs=None, fp=io.BytesIO(b'{"error":"unauthorized"}'))


class BillingUsecaseFlowTest(unittest.TestCase):
    def setUp(self) -> None:
        self.temp_dir = tempfile.mkdtemp(prefix="billing_usecase_test_")
        self.upload_dir = os.path.join(self.temp_dir, "data")
        self.output_dir = os.path.join(self.temp_dir, "output")
        os.makedirs(self.upload_dir, exist_ok=True)
        os.makedirs(self.output_dir, exist_ok=True)

        # Point runtime settings to temporary test folders.
        settings.upload_dir = self.upload_dir
        settings.output_dir = self.output_dir

        # Provide cost center mapping used by step-6 in the use case.
        with open(os.path.join(self.upload_dir, "cost_centers.csv"), "w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=["myid", "cost_center"])
            writer.writeheader()
            writer.writerow({"myid": "M100", "cost_center": "CC-100"})
            writer.writerow({"myid": "M200", "cost_center": "CC-200"})

        self.client = app.test_client()

    def tearDown(self) -> None:
        shutil.rmtree(self.temp_dir, ignore_errors=True)

    def test_upload_end_to_end_usecase(self) -> None:
        csv_data = "\n".join(
            [
                "amount,user_type,region,country,holidex,person_holidex,course_name,myid",
                "0,C,AMER,United States of America,AB123,AB123,Excel, M100",
                "120,C,AMER,Canada,XX111,ZX9K1,Data Science (Core)&AI-101!!!,M100",
                "220,F,AMER,Mexico,MX111,MX111,Franchise$Billing,M200",
                "340,H,AMEA,India,AM222,AM222,Leadership@@Workshop,M300",
                "150,C,GC,Australia,GC333,GC333,GC-Program (Intro),M200",
            ]
        )

        response = self.client.post(
            "/upload",
            data={"file": (io.BytesIO(csv_data.encode("utf-8")), "billing_input.csv")},
            content_type="multipart/form-data",
        )

        self.assertEqual(response.status_code, 200, response.get_data(as_text=True))

        payload = response.get_json()
        assert payload is not None

        self.assertEqual(payload["total_rows"], 5)
        self.assertEqual(payload["zero_rows"], 1)
        self.assertEqual(payload["corp_rows"], 2)
        self.assertEqual(payload["non_corp_rows"], 2)

        zero_path = payload["zero_data_path"]
        corp_path = payload["corp_paid_path"]
        non_corp_path = payload["non_corp_paid_path"]

        self.assertTrue(Path(zero_path).exists(), zero_path)
        self.assertTrue(Path(corp_path).exists(), corp_path)
        self.assertTrue(Path(non_corp_path).exists(), non_corp_path)

        zero_wb = load_workbook(zero_path)
        self.assertIn("ZERO_DATA", zero_wb.sheetnames)

        corp_wb = load_workbook(corp_path)
        for required_sheet in ["SUMMARY", "ALL", "AMER", "MEXICO", "AMEA", "EMEAA", "GC"]:
            self.assertIn(required_sheet, corp_wb.sheetnames)

        non_corp_wb = load_workbook(non_corp_path)
        for required_sheet in ["SUMMARY", "ALL", "AMER", "MEXICO", "AMEA", "EMEAA", "GC"]:
            self.assertIn(required_sheet, non_corp_wb.sheetnames)

        corp_all = _sheet_records(corp_path, "ALL")
        non_corp_all = _sheet_records(non_corp_path, "ALL")

        self.assertEqual(len(corp_all), 2)
        self.assertEqual(len(non_corp_all), 2)

        # Holidex should be synced from valid person_holidex for mismatched row.
        holidex_values = {str(row.get("holidex")) for row in corp_all}
        self.assertIn("ZX9K1", holidex_values)

        # Cost center should be filled for mapped MYIDs.
        cost_centers = {str(row.get("cost_center")) for row in corp_all}
        self.assertTrue("CC-100" in cost_centers or "CC-200" in cost_centers)

        # Course name cleanup should remove disallowed special characters.
        cleaned_course_names = [str(row.get("course_name")) for row in corp_all + non_corp_all]
        self.assertTrue(any("!!!" not in name and "@@" not in name and "$" not in name for name in cleaned_course_names))


class SharePointClientConfigurationTest(unittest.TestCase):
    def test_download_client_reports_missing_site_configuration(self) -> None:
        client = SharePointDownloadClient(
            tenant_id="tenant",
            client_id="client",
            client_secret="secret",
            site_url="",
            site_id="",
        )
        client._site_url = ""
        client._site_id = ""

        with patch(
            "app.services.sharepoint_download_service.urlopen",
            side_effect=_unauthorized_http_error("https://graph.microsoft.com/v1.0/sites/root"),
        ):
            with self.assertRaisesRegex(ValueError, "SHAREPOINT_SITE_URL or SHAREPOINT_SITE_ID"):
                client._get_site_id("token")

    def test_upload_client_reports_missing_site_configuration(self) -> None:
        client = SharePointUploadClient(
            tenant_id="tenant",
            client_id="client",
            client_secret="secret",
            site_url="",
            site_id="",
        )
        client._site_url = ""
        client._site_id = ""

        with patch(
            "app.services.sharepoint_upload_service.urlopen",
            side_effect=_unauthorized_http_error("https://graph.microsoft.com/v1.0/sites/root"),
        ):
            with self.assertRaisesRegex(ValueError, "SHAREPOINT_SITE_URL or SHAREPOINT_SITE_ID"):
                client._get_site_id("token")

    def test_download_client_retries_drive_lookup_with_password_auth(self) -> None:
        client = SharePointDownloadClient(
            tenant_id="tenant",
            client_id="client",
            client_secret="secret",
            username="user@example.com",
            password="password",
            site_id="site-id",
        )
        client._token = "client-token"
        client._auth_mode = "client_credentials"
        client._token_expires_at = datetime.now(timezone.utc) + timedelta(minutes=5)

        def fake_urlopen(request, timeout=30):
            auth_header = request.headers.get("Authorization")
            if request.full_url.endswith("/drives") and auth_header == "Bearer client-token":
                raise _unauthorized_http_error(request.full_url)
            if request.full_url.endswith("/token"):
                return io.BytesIO(b'{"access_token": "password-token", "expires_in": 3600}')
            if request.full_url.endswith("/drives") and auth_header == "Bearer password-token":
                return io.BytesIO(b'{"value": [{"name": "Documents", "id": "drive-id"}]}')
            raise AssertionError(f"Unexpected request: {request.full_url} {auth_header}")

        with patch("app.services.sharepoint_download_service.urlopen", side_effect=fake_urlopen):
            self.assertEqual(client._get_drive_id("client-token", "site-id"), "drive-id")


if __name__ == "__main__":
    unittest.main(verbosity=2)
